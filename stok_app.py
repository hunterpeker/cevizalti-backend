from pdf_excel import treeview_pdf_aktar, treeview_excel_aktar
from data import yukle, kaydet
from reports import (rapor_ekrani,satis_raporu,stok_raporu,kullanici_raporu,odeme_raporu)
from reports import urun_satis_adet_raporu, satin_alma_fiyat_raporu
from reports import kar_zarar_yeni
from services.backend_client import (
    siparis_gonder,
    adisyonlari_yukle,
    masa_kapat as backend_masa_kapat,
    siparis_iptal,
    masa_tasi,
    gelirleri_al
)
from core.update_service import dosya_indir, github_version_bilgisi_al
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from PIL import Image, ImageTk
from tkcalendar import DateEntry
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import json
import os
import hashlib
from datetime import datetime
import shutil
from pystray import Icon, Menu, MenuItem
import win32print
import ui_theme
FONT_SMALL  = ui_theme.FONT_SMALL
FONT_NORMAL = ui_theme.FONT_NORMAL
FONT_BIG    = ui_theme.FONT_BIG
FONT_TITLE  = ui_theme.FONT_TITLE
import win32ui
import socket
import sys

APP_VERSION = "1.0.0"

GITHUB_REPO = "hunterpeker/cevizalti-backend"
GITHUB_BRANCH = "main"

GITHUB_RAW_BASE = (
    f"https://raw.githubusercontent.com/{GITHUB_REPO}/{GITHUB_BRANCH}"
)

GUNCELLENECEK_DOSYALAR = [
    "app.py",
    "data.py",
    "pdf_excel.py",
    "reports.py",
    "stok_app.py",
    "ui_theme.py",
]
toplam_tutar = 0.0
indirim_orani = 0.0
indirimli_toplam = 0.0
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MENU_YAPISI = {
    "Restoran": {
        "Dashboard": "dashboard",
        "Menü Yönetimi": "menu_yonetimi",
        "Adisyonlar": "adisyon_ekrani"
    },
    "Depo": {
        "Stok Görüntüle": "stok_goster",
        "Stok Çıkış": "stok_cikis",
        "Stok Hareketleri": "stok_hareketleri",
        "Ürün Tanımlama": "urun_yonetimi"
    },
    "Satın Alma": {
        "Gider Girişi": "gider_ekle"
    },
    "Muhasebe": {
        "Gelir Girişi": "manuel_gelir_ekle",
        "Kâr / Zarar": "kar_zarar"
    },
    "Rapor": {
        "Satış Raporu": "satis_raporu",
        "Ödeme Raporu": "odeme_raporu",
        "Ürün Satış Raporu": "urun_satis_adet_raporu",
        "Satın Alma Fiyat Raporu": "satin_alma_fiyat_raporu",
         "Kâr / Zarar (Yeni)": "kar_zarar_yeni"
    },
    "Admin": {
        "Kullanıcı Yönetimi": "kullanici_yonetimi"
    },
    "Yedekleme": {
        "Yedek Al": "yedek_al",
        "Yedekten Yükle": "yedekten_yukle"
    }
}
def varsayilan_yetkiler():
    return {
        "Restoran": {
            "Dashboard": False,
            "Menü Yönetimi": False,
            "Adisyonlar": False
        },
        "Depo": {
            "Stok Görüntüle": False,
            "Stok Çıkış": False,
            "Stok Hareketleri": False,
            "Ürün Tanımlama": False
        },
        "Satın Alma": {
            "Gider Girişi": False
        },
        "Muhasebe": {
            "Gelir Girişi": False,
            "Kâr / Zarar": False
        },
        "Rapor": {
            "Satış Raporu": False,
            "Ödeme Raporu": False,
            "Kullanıcı Raporu": False,
            "Kâr / Zarar (Yeni)": False
        },
        "Admin": {
            "Kullanıcı Yönetimi": False
        },
        "Yedekleme": {
            "Yedek Al": False,
            "Yedekten Geri Yükle": False
        }
    }

# ================= GARDON SERVER (ARKA PLAN) =================
SERVER_DIR = os.path.join(BASE_DIR, "garson_server")
PYTHON_EXE = os.path.join(SERVER_DIR, "venv", "Scripts", "python.exe")
SERVER_APP = os.path.join(SERVER_DIR, "app.py")
ICON_PATH = os.path.join(BASE_DIR, "ixq.ico")

# ================= DOSYALAR =================
DOSYA_URUN = os.path.join(BASE_DIR, "products.json")
DOSYA_HAREKET = os.path.join(BASE_DIR, "hareketler.json")
DOSYA_GIDER = os.path.join(BASE_DIR, "expenses.json")
DOSYA_GELIR = os.path.join(BASE_DIR, "revenue.json")
DOSYA_USER = os.path.join(BASE_DIR, "users.json")
DOSYA_MENU = os.path.join(BASE_DIR, "menu.json")
DOSYA_SATIS_DETAY = os.path.join(BASE_DIR, "satis_detay.json")
TOPLAM_MASA = 30

# ================= GLOBAL FONT & ÖLÇEK =================
import tkinter.font as tkfont

APP_FONT = "Segoe UI"

FONT_SMALL  = (APP_FONT, 10)
FONT_NORMAL = (APP_FONT, 11)
FONT_BIG    = (APP_FONT, 13)
FONT_TITLE  = (APP_FONT, 15, "bold")


# ================= YARDIMCI =================
def yetkisi_var(mi_ana, mi_alt=None):
    try:
        y = kullanicilar[aktif_kullanici]["yetkiler"]
        if mi_alt is None:
            return any(y.get(mi_ana, {}).values())
        return y.get(mi_ana, {}).get(mi_alt, False)
    except:
        return False

def temizle(parent):
    for w in parent.winfo_children():
        w.destroy()

from tkinter import filedialog
import shutil

def pdf_kaydet_dialog(gecici_pdf, masa):
    varsayilan_ad = f"Adisyon_{masa}_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.pdf"

    hedef = filedialog.asksaveasfilename(
        title="Adisyonu Kaydet",
        defaultextension=".pdf",
        initialfile=varsayilan_ad,
        filetypes=[("PDF Dosyası", "*.pdf")]
    )

    if not hedef:
        return False

    shutil.copy(gecici_pdf, hedef)
    return True
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import tempfile
import os
from datetime import datetime

pdfmetrics.registerFont(UnicodeCIDFont("HeiseiMin-W3"))

def yazici_sec_dialog():
    win = tk.Toplevel()
    win.title("Yazdırma Türü Seç")
    win.geometry("300x200")
    win.grab_set()
    win.resizable(False, False)

    secim = tk.StringVar(value="A4")
    sonuc = {"val": A4}

    ttk.Label(
        win,
        text="Yazdırma Türünü Seçin",
        font=("Segoe UI", 11, "bold")
    ).pack(pady=12)

    ttk.Radiobutton(
        win,
        text="🧾 A4 Yazıcı",
        variable=secim,
        value="A4"
    ).pack(anchor="w", padx=30, pady=4)

    ttk.Radiobutton(
        win,
        text="🖨 Termal (80mm)",
        variable=secim,
        value="TERMAL"
    ).pack(anchor="w", padx=30, pady=4)

    def onayla():
        sonuc["val"] = secim.get()
        win.destroy()

    ttk.Button(
        win,
        text="Yazdır",
        width=16,
        command=onayla
    ).pack(pady=18)

    win.wait_window()
    return sonuc["val"]

def pdf_fis_olustur(masa, urunler, toplam, odeme, kullanici):
    # ===== TARİH / SAAT =====
    now = datetime.now()
    gun = now.strftime("%Y-%m-%d")
    saat = now.strftime("%H-%M-%S")

    # ===== MASAÜSTÜ / FİŞLER / TARİH =====
    base_dir = os.path.join(
        os.path.expanduser("~/Desktop"),
        "Fisler",
        gun
    )
    os.makedirs(base_dir, exist_ok=True)

    # ===== DOSYA YOLU =====
    dosya = os.path.join(
        base_dir,
        f"adisyon_{masa.replace(' ', '_')}_{saat}.pdf"
    )

    # ===== PDF =====
    c = canvas.Canvas(dosya, pagesize=A4)
    w, h = A4
    y = h - 60

    def satir(txt, size=11):
        nonlocal y
        c.setFont("HeiseiMin-W3", size)
        c.drawCentredString(w / 2, y, txt)
        y -= size + 6

    satir("CEVİZ ALTI RESTAURANT", 16)
    satir("Adisyon Fişi", 12)
    satir("-" * 40, 10)

    satir(f"Masa: {masa}")
    satir(f"Ödeme: {odeme}")
    satir(f"Garson: {kullanici}")
    satir(now.strftime("%d.%m.%Y %H:%M"))
    satir("-" * 40, 10)

    for u in urunler:
        satir(f"{u['ad']}  x{u['adet']}  {u['tutar']:.2f} ₺", 11)

    satir("-" * 40, 10)
    satir(f"TOPLAM: {toplam:.2f} ₺", 14)

    c.showPage()
    c.save()

    return dosya

# ================= FİŞ YAZDIRMA =================
def fis_yazdir(masa, urunler, toplam, odeme, kullanici):
    printer_name = win32print.GetDefaultPrinter()
    hprinter = win32ui.CreateDC()
    hprinter.CreatePrinterDC(printer_name)

    hprinter.StartDoc("Adisyon Fişi")
    hprinter.StartPage()

    y = 100

    def satir(text):
        nonlocal y
        hprinter.TextOut(100, y, text)
        y += 30

    satir("CEVİZ ALTI RESTAURANT")
    satir("---------------------------")
    satir(f"Masa   : {masa}")
    satir(f"Ödeme : {odeme}")
    satir(f"Garson: {kullanici}")
    satir("---------------------------")

    for u in urunler:
        satir(f"{u['ad']} x{u['adet']}  {u['tutar']:.2f} ₺")

    satir("---------------------------")
    satir(f"TOPLAM: {toplam:.2f} ₺")
    satir(datetime.now().strftime("%d.%m.%Y %H:%M"))

    hprinter.EndPage()
    hprinter.EndDoc()
    hprinter.DeleteDC()

def hashle(s):
    return hashlib.sha256(s.encode()).hexdigest()

from tkinter import messagebox

def sistem_guncelle_onay():
    cevap = messagebox.askyesno(
        "Sistem Güncelleme",
        "⚠️ DİKKAT!\n\n"
        "Bu işlem sistem dosyalarını günceller.\n"
        "Devam etmek istiyor musunuz?"
    )

    if not cevap:
        return

    guncelleme_baslat()

def guncelleme_baslat():
    github_data = github_version_bilgisi_al()

    if not github_data:
        messagebox.showerror(
            "Güncelleme Kontrolü",
            "GitHub'a bağlanılamadı.\nİnternet bağlantısını kontrol edin."
        )
        return

    remote_version = github_data.get("version", "bilinmiyor")
    remote_date = github_data.get("date", "")
    remote_desc = github_data.get("desc", "")

    if remote_version == APP_VERSION:
        messagebox.showinfo(
            "Sistem Güncel",
            f"Sürüm: {APP_VERSION}\n\nProgram zaten güncel."
        )
        return

    # 🔴 ONAY EKRANI (KRİTİK KISIM)
    onay = messagebox.askyesno(
        "⚠️ Yeni Güncelleme Var",
        f"Mevcut Sürüm : {APP_VERSION}\n"
        f"Yeni Sürüm   : {remote_version}\n\n"
        f"Tarih: {remote_date}\n"
        f"Açıklama:\n{remote_desc}\n\n"
        "Güncelleme yapılmadan önce\n"
        "yedek alındığından emin olun.\n\n"
        "Güncellemeye devam edilsin mi?"
    )

    if not onay:
        return

    # ✅ ONAYDAN SONRA GERÇEK GÜNCELLEME
    gercek_guncelleme_baslat(github_data)


def gercek_guncelleme_baslat(github_data):
    try:
        # 1️⃣ YEDEK
        yedek_al()

        # 2️⃣ DOSYALARI İNDİR
        for dosya in GUNCELLENECEK_DOSYALAR:
            url = f"{GITHUB_RAW_BASE}/{dosya}"
            hedef = os.path.join(os.getcwd(), dosya)

            icerik = dosya_indir(url, dosya)

            with open(hedef, "wb") as f:
                f.write(icerik)

        messagebox.showinfo(
            "Güncelleme Tamamlandı",
            "Güncelleme başarıyla yapıldı.\n\n"
            "Program şimdi yeniden başlatılacak."
        )

        # 3️⃣ PROGRAMI YENİDEN BAŞLAT
        python = sys.executable
        os.execl(python, python, *sys.argv)

    except Exception as e:
        messagebox.showerror(
            "Güncelleme Hatası",
            f"Güncelleme sırasında hata oluştu:\n\n{e}"
        )


def turkce_key(s):
    if not isinstance(s, str):
        return s

    cevir = str.maketrans(
        "ÇĞİÖŞÜçğıöşü",
        "CGIOSUcgiosu"
    )
    return s.translate(cevir).lower()

def treeview_sirala(tree, col, ters, tip="str"):
    veri = []

    for k in tree.get_children(""):
        deger = tree.set(k, col)

        try:
            if tip == "float":
                deger = float(str(deger).replace(",", "").replace("+", ""))
            elif tip == "int":
                deger = int(deger)
            else:
                deger = turkce_key(deger)
        except:
            pass

        veri.append((deger, k))

    veri.sort(reverse=ters)

    for index, (_, k) in enumerate(veri):
        tree.move(k, "", index)

    tree.heading(
        col,
        command=lambda: treeview_sirala(tree, col, not ters, tip)
    )
def recete_stok_kontrol_ve_dus(menu_id, adet):
    """
    Menü satışı sırasında reçeteye göre stok kontrolü yapar ve stok düşer.
    Stok yetersizse False döner.
    """

    menu = menuler[menu_id]

    # 1️⃣ Önce KONTROL
    for r in menu["recete"]:
        urun = urunler[r["urun_kod"]]
        gereken = r["miktar"] * adet

        if urun["stok"] < gereken:
            messagebox.showerror(
                "Yetersiz Stok",
                f"{urun['ad']} için stok yetersiz!\n"
                f"Gereken: {gereken} {urun['birim']}\n"
                f"Mevcut: {urun['stok']} {urun['birim']}"
            )
            return False

    # 2️⃣ STOK DÜŞ
    for r in menu["recete"]:
        urun = urunler[r["urun_kod"]]
        dusulecek = r["miktar"] * adet

        urun["stok"] = round(urun["stok"] - dusulecek, 2)
        hareketler.append({
            "tarih": datetime.now().strftime("%d.%m.%Y %H:%M"),
            "kullanici": aktif_kullanici,
            "urun": urun["ad"],
            "miktar": dusulecek,
            "tur": "ÇIKIŞ"
        })

    # 3️⃣ KAYDET
    kaydet(DOSYA_URUN, urunler)
    kaydet(DOSYA_HAREKET, hareketler)

    return True

# ================= KULLANICI =================
if not os.path.exists(DOSYA_USER):
    kaydet(
        DOSYA_USER,
        {
            "Admin": {
                "password": hashle("1234"),
                "role": "admin",
                "aktif": True,
                "yetkiler": varsayilan_yetkiler()
            }
        }
    )

kullanicilar = yukle(DOSYA_USER, {})

# ===== ESKİ KULLANICILAR İÇİN YETKİ TAMAMLAMA =====
degisti = False

for ad, veri in kullanicilar.items():
    if "yetkiler" not in veri:
        veri["yetkiler"] = varsayilan_yetkiler()
        degisti = True

    if "aktif" not in veri:
        veri["aktif"] = True
        degisti = True

if degisti:
    kaydet(DOSYA_USER, kullanicilar)

aktif_kullanici = "Admin"
aktif_rol = "admin"

# ================= ROOT =================
root = tk.Tk()
def uygula_antrasit_tema(root):
    style = ttk.Style(root)
    style.theme_use("default")

    ANA_BG   = "#2b2b2b"
    IKINCIL  = "#3c3f41"
    YAZI     = "#e6e6e6"
    VURGU    = "#4e5254"

    root.configure(bg=ANA_BG)

    style.configure(
        ".",
        background=ANA_BG,
        foreground=YAZI,
        fieldbackground=IKINCIL,
        bordercolor=VURGU,
        font=("Segoe UI", 10)
    )

    style.configure(
        "TFrame",
        background=ANA_BG
    )

    style.configure(
        "TLabel",
        background=ANA_BG,
        foreground=YAZI
    )

    style.configure(
        "TButton",
        background=IKINCIL,
        foreground=YAZI,
        padding=6
    )

    style.map(
        "TButton",
        background=[("active", "#505355")]
    )

    style.configure(
        "Treeview",
        background=IKINCIL,
        fieldbackground=IKINCIL,
        foreground=YAZI,
        rowheight=28
    )

    style.configure(
        "Treeview.Heading",
        background="#1e1e1e",
        foreground=YAZI,
        font=("Segoe UI", 10, "bold")
    )

    style.map(
        "Treeview",
        background=[("selected", "#606366")]
    )

icon_path = os.path.join(BASE_DIR, "ceviz.ico")
root.iconbitmap(icon_path)

# ================= GLOBAL TTK STYLE =================
style = ttk.Style()
style.theme_use("default")

# 🔹 TÜM TTK WIDGET FONT
style.configure(
    ".",
    font=("Segoe UI", 11)
)

# 🔹 TREEVIEW SATIRLARI
style.configure(
    "Treeview",
    font=("Segoe UI", 11),
    rowheight=30,
    background="#ffffff",
    fieldbackground="#ffffff"
)

# 🔹 TREEVIEW BAŞLIKLARI
style.configure(
    "Treeview.Heading",
    font=("Segoe UI", 13, "bold"),
    background="#e0e0e0",
    foreground="#000000",
    relief="raised"
)

# 🔹 AKTİF BAŞLIK RENGİ
style.map(
    "Treeview.Heading",
    background=[("active", "#d6d6d6")]
)

# ================= DPI + COMBOBOX POPDOWN FIX =================
root.update_idletasks()
root.tk.call("tk", "scaling", 1.25)  # %125 Windows için ideal

root.option_add("*TCombobox*Listbox.font", ("Segoe UI", 14))
root.option_add("*TCombobox*Listbox.selectBackground", "#3498db")
root.option_add("*TCombobox*Listbox.selectForeground", "white")

# ================= MENÜ FONT FIX =================
MENU_FONT = ("Segoe UI", 12)

root.option_add("*Menu.font", MENU_FONT)

# ================= DPI / ÖLÇEK SABİTLE =================
root.update_idletasks()

# ================= GLOBAL STYLE =================

# Tüm ttk widget’lar

root.title("Ceviz Altı Restaurant")
root.state("zoomed")

# ================= ANA LAYOUT =================
# 🔹 ÜST MENÜ (zaten aşağıda bağlanacak)
# (menubar koduna dokunmuyoruz)
# 🔹 ORTA ALAN (TÜM EKRANLAR BURADA AÇILACAK)
content_frame = tk.Frame(root, bg="white")
content_frame.pack(fill="both", expand=True)

# 🔹 ALT DURUM ÇUBUĞU (SAAT / TARİH)
status_bar = ttk.Frame(root)
status_bar.pack(fill="x", side="bottom")

saat_lbl = ttk.Label(status_bar, anchor="e")
saat_lbl.pack(fill="x", padx=10)

def saat_guncelle():
    saat_lbl.config(text=datetime.now().strftime(" %d.%m.%Y  |  %H:%M:%S"))
    root.after(1000, saat_guncelle)

saat_guncelle()

# ================= ORTA ALAN TEMİZLE =================
def temizle_orta_alan():
    # content_frame içini tamamen temizle
    for w in content_frame.winfo_children():
        w.destroy()

    # layout reset
    content_frame.pack_forget()
    content_frame.pack(fill="both", expand=True)


def esc_handler(event=None):
    try:
        w = root.focus_get()
        if w:
            top = w.winfo_toplevel()
            if top != root:
                top.destroy()
                return "break"
    except:
        pass

    ana_sayfa_goster()
    return "break"

# ⬇⬇⬇ EN SONDA ⬇⬇⬇
root.bind_all("<Escape>", esc_handler)

# ================= LOGIN =================

def login_ekrani():
    global aktif_kullanici, aktif_rol

    temizle_orta_alan()

    frame = ttk.Frame(content_frame, padding=40)
    frame.pack(expand=True)

    kullanici_adi = tk.StringVar()
    sifre = tk.StringVar()

    def giris_yap():
        nonlocal kullanici_adi, sifre
        global aktif_kullanici, aktif_rol

        k = kullanici_adi.get().strip()
        s = sifre.get()

        if not k or not s:
            messagebox.showerror("Hata", "Alanlar boş")
            return

        if k not in kullanicilar:
            messagebox.showerror("Hata", "Kullanıcı bulunamadı")
            return

        if kullanicilar[k]["password"] != hashle(s):
            messagebox.showerror("Hata", "Şifre hatalı")
            return

        if not kullanicilar[k].get("aktif", True):
            messagebox.showerror("Yetkisiz", "Bu kullanıcı pasif")
            return

        aktif_kullanici = k
        aktif_rol = kullanicilar[k]["role"]
        ana_ekran()
        kritik_kontrol()

    ttk.Label(frame, text="Kullanıcı Adı").pack()
    ttk.Entry(frame, textvariable=kullanici_adi).pack()

    ttk.Label(frame, text="Şifre").pack()
    sifre_entry = ttk.Entry(frame, textvariable=sifre, show="*")
    sifre_entry.pack()

    ttk.Button(frame, text="Giriş", command=giris_yap).pack(pady=20)

    # ENTER ile giriş
    sifre_entry.bind("<Return>", lambda e: giris_yap())


def menu_yonetimi():
    temizle_orta_alan()

    ana = ttk.Frame(content_frame, padding=20)
    ana.pack(fill="both", expand=True)

    ttk.Label(ana, text="🍽 Menü Yönetimi", font=FONT_TITLE).pack(pady=(0, 10))

    govde = ttk.Frame(ana)
    govde.pack(fill="both", expand=True)

    secili_menu = {"id": None}

    # ================= FONKSİYONLAR =================

    def menu_yukle():
        menu_tree.delete(*menu_tree.get_children())
        for mid, m in menuler.items():
            menu_tree.insert("", "end", iid=mid, values=(m["ad"], f'{m["fiyat"]:.2f}'))

    def recete_goster(event=None):
        sec = menu_tree.focus()
        if not sec:
            return
        secili_menu["id"] = sec
        recete_tree.delete(*recete_tree.get_children())
        for r in menuler[sec]["recete"]:
            u = urunler[r["urun_kod"]]
            recete_tree.insert("", "end", values=(u["ad"], u["birim"], r["miktar"]))

    def menu_ekle():
        p = tk.Toplevel(root)
        p.title("Menü Ekle")
        p.geometry("300x260")
        p.grab_set()

        ad = tk.StringVar()
        fiyat = tk.StringVar()
        mutfak = tk.BooleanVar(value=True)  # 👈 YENİ

        ttk.Label(p, text="Menü Adı").pack(pady=5)
        ttk.Entry(p, textvariable=ad).pack()

        ttk.Label(p, text="Fiyat").pack(pady=5)
        ttk.Entry(p, textvariable=fiyat).pack()

        ttk.Checkbutton(
            p,
            text="🍳 Mutfağa gönderilsin",
            variable=mutfak
        ).pack(pady=10)

        def kaydet_menu():
            try:
                f = float(fiyat.get())
            except:
                messagebox.showerror("Hata", "Fiyat geçersiz")
                return

            yeni_id = str(max(map(int, menuler.keys()), default=0) + 1)

            menuler[yeni_id] = {
                "ad": ad.get().strip(),
                "fiyat": f,
                "recete": [],
                "mutfak": mutfak.get()   # 👈 KRİTİK SATIR
            }

            kaydet(DOSYA_MENU, menuler)
            menu_yukle()
            p.destroy()

        ttk.Button(p, text="Kaydet", command=kaydet_menu).pack(pady=10)


    def menu_duzenle():
        sec = menu_tree.focus()
        if not sec:
            return

        menu = menuler[sec]

        p = tk.Toplevel(root)
        p.title("Menü Düzenle")
        p.geometry("300x260")
        p.grab_set()

        ad = tk.StringVar(value=menu["ad"])
        fiyat = tk.StringVar(value=str(menu["fiyat"]))
        mutfak = tk.BooleanVar(value=menu.get("mutfak", True))  # ✅ DOĞRU YER

        ttk.Label(p, text="Menü Adı").pack(pady=5)
        ttk.Entry(p, textvariable=ad).pack()

        ttk.Label(p, text="Fiyat").pack(pady=5)
        ttk.Entry(p, textvariable=fiyat).pack()

        ttk.Checkbutton(
            p,
            text="🍳 Mutfağa gönderilsin",
            variable=mutfak
        ).pack(pady=10)

        def kaydet_deg():
            try:
                f = float(fiyat.get())
            except:
                messagebox.showerror("Hata", "Fiyat geçersiz")
                return

            menu["ad"] = ad.get().strip()
            menu["fiyat"] = f
            menu["mutfak"] = mutfak.get()   # ✅ KRİTİK SATIR

            kaydet(DOSYA_MENU, menuler)
            menu_yukle()
            p.destroy()

        ttk.Button(p, text="Kaydet", command=kaydet_deg).pack(pady=10)


    def menu_sil():
        sec = menu_tree.focus()
        if not sec:
            return
        if not messagebox.askyesno("Sil", "Menü silinsin mi?"):
            return
        menuler.pop(sec)
        kaydet(DOSYA_MENU, menuler)
        menu_yukle()
        recete_tree.delete(*recete_tree.get_children())

    def recete_ekle():
        if not secili_menu["id"]:
            return

        p = tk.Toplevel(root)
        p.title("Reçeteye Ürün")
        p.geometry("300x200")
        p.grab_set()

        urun = tk.StringVar()
        miktar = tk.StringVar()

        ttk.Combobox(p, values=[u["ad"] for u in urunler.values()],
                     textvariable=urun, state="readonly").pack(pady=5)
        ttk.Entry(p, textvariable=miktar).pack(pady=5)

        def kaydet_recete():
            try:
                m = float(miktar.get())
            except:
                return
            kod = next(k for k, v in urunler.items() if v["ad"] == urun.get())
            menuler[secili_menu["id"]]["recete"].append({"urun_kod": kod, "miktar": m})
            kaydet(DOSYA_MENU, menuler)
            recete_goster()
            p.destroy()

        ttk.Button(p, text="Kaydet", command=kaydet_recete).pack(pady=10)

    def recete_sil():
        sec = recete_tree.focus()
        if not sec:
            return
        urun_adi = recete_tree.item(sec)["values"][0]
        recete = menuler[secili_menu["id"]]["recete"]
        recete[:] = [r for r in recete if urunler[r["urun_kod"]]["ad"] != urun_adi]
        kaydet(DOSYA_MENU, menuler)
        recete_goster()

    def recete_duzenle():
        sec = recete_tree.focus()
        if not sec:
            return

        urun_adi, birim, eski_miktar = recete_tree.item(sec)["values"]

        p = tk.Toplevel(root)
        p.title("Miktar Düzenle")
        p.geometry("250x150")
        p.grab_set()

        miktar = tk.StringVar(value=str(eski_miktar))

        ttk.Label(p, text=f"{urun_adi} ({birim})").pack(pady=5)
        ttk.Entry(p, textvariable=miktar).pack(pady=5)

        def kaydet_miktar():
            try:
                yeni_miktar = float(miktar.get())
            except:
                messagebox.showerror("Hata", "Geçersiz miktar")
                return

            # 🔴 VERİYİ GÜNCELLE
            recete = menuler[secili_menu["id"]]["recete"]
            for r in recete:
                if urunler[r["urun_kod"]]["ad"] == urun_adi:
                    r["miktar"] = yeni_miktar
                    break

            kaydet(DOSYA_MENU, menuler)

            # 🔴 EKRANI ANINDA YENİLE
            recete_goster()

            # 🔴 PENCEREYİ KAPAT
            p.destroy()

        ttk.Button(p, text="Kaydet", command=kaydet_miktar).pack(pady=10)



    # ================= SOL: MENÜLER =================
    sol = ttk.LabelFrame(govde, text="Menüler", padding=10)
    sol.pack(side="left", fill="both", expand=False, padx=(0, 15))

    # ---- MENÜ LİSTESİ ----
    menu_liste_frame = ttk.Frame(sol)
    menu_liste_frame.pack(fill="both", expand=True)

    menu_tree = ttk.Treeview(
        menu_liste_frame,
        columns=("Ad", "Fl"),
        show="headings"
    )
    menu_tree.heading("Ad", text="Menü Adı")
    menu_tree.heading("Fl", text="Fiyat")
    menu_tree.column("Ad", width=220, anchor="w")
    menu_tree.column("Fl", width=90, anchor="e")
    menu_tree.pack(side="left", fill="both", expand=True)
    menu_tree.bind("<<TreeviewSelect>>", recete_goster)
    menu_scroll = ttk.Scrollbar(
        menu_liste_frame, orient="vertical", command=menu_tree.yview
    )
    menu_scroll.pack(side="right", fill="y")
    menu_tree.configure(yscrollcommand=menu_scroll.set)

    # ---- MENÜ BUTONLARI (HER ZAMAN GÖRÜNÜR) ----
    menu_btn = ttk.Frame(sol)
    menu_btn.pack(fill="x", pady=(8, 0))

    ttk.Button(
        menu_btn, text="➕ Menü Ekle", command=menu_ekle
    ).grid(row=0, column=0, sticky="ew", padx=4)

    ttk.Button(
        menu_btn, text="✏️ Menü Düzenle", command=menu_duzenle
    ).grid(row=0, column=1, sticky="ew", padx=4)

    ttk.Button(
        menu_btn, text="🗑 Menü Sil", command=menu_sil
    ).grid(row=0, column=2, sticky="ew", padx=4)

    menu_btn.columnconfigure((0, 1, 2), weight=1)


    # ================= SAĞ =================
    sag = ttk.LabelFrame(govde, text="Reçete", padding=10)
    sag.pack(side="left", fill="both", expand=True)

    recete_tree = ttk.Treeview(sag, columns=("Ürün", "Birim", "Miktar"), show="headings")
    for c in ("Ürün", "Birim", "Miktar"):
        recete_tree.heading(c, text=c)
    recete_tree.pack(fill="both", expand=True)

    btn_sag = ttk.Frame(sag)
    btn_sag.pack(fill="x", pady=8)

    ttk.Button(btn_sag, text="➕ Reçeteye Ürün", command=recete_ekle)\
        .pack(side="left", expand=True, fill="x", padx=4)

    ttk.Button(btn_sag, text="✏️ Miktar Düzenle", command=recete_duzenle)\
        .pack(side="left", expand=True, fill="x", padx=4)

    ttk.Button(btn_sag, text="🗑 Ürünü Sil", command=recete_sil)\
        .pack(side="left", expand=True, fill="x", padx=4)


    menu_yukle()


def adisyon_ekrani():
    adisyonlar = adisyonlari_yukle()
    secili_masa = {"ad": None}
    temizle_orta_alan()

    # ================= MASA SAĞ TIK MENÜSÜ =================
    masa_menu = tk.Menu(root, tearoff=0)

    def masa_tasi_dialog(eski_masa):
        p = tk.Toplevel(root)
        p.title("Masayı Taşı")
        p.geometry("300x180")
        p.grab_set()

        ttk.Label(p, text=f"{eski_masa} →").pack(pady=10)

        hedef = tk.StringVar()

        bos_masalar = [
            m for m in masa_butonlari.keys()
            if m != eski_masa and m not in adisyonlar
        ]

        cb = ttk.Combobox(
            p,
            textvariable=hedef,
            values=bos_masalar,
            state="readonly",
            width=20
        )
        cb.pack(pady=10)

        def onayla():
            if not hedef.get():
                return

            masa_tasi(eski_masa, hedef.get(), aktif_kullanici)

            p.destroy()

            # 🔄 ADİSYONLARI YENİLE
            adisyonlar.clear()
            adisyonlar.update(adisyonlari_yukle())
            masa_renk_guncelle()

        ttk.Button(p, text="Taşı", command=onayla).pack(pady=15)


    # ================= VERİ =================
    adisyonlar = adisyonlari_yukle()
    masalar = [f"Masa {i}" for i in range(1, TOPLAM_MASA + 1)] + ["Paket", "Gel-Al"]
    masa_butonlari = {}

    # ================= ANA FRAME =================
    ana = ttk.Frame(content_frame, padding=20)
    ana.pack(fill="both", expand=True)

    # ================= SOL : MASALAR =================
    sol = ttk.Frame(ana, width=300)
    sol.pack(side="left", fill="y", padx=(0, 20))

    sol.columnconfigure(0, weight=1)
    sol.columnconfigure(1, weight=1)

    ttk.Label(
        sol,
        text="Masalar",
        font=("Segoe UI", 14, "bold")
    ).grid(row=0, column=0, columnspan=2, pady=(0, 10))

    # ================= SAĞ : ADİSYON =================
    sag = ttk.Frame(ana)
    sag.pack(side="left", fill="both", expand=True)

    baslik = ttk.Label(
        sag,
        text="Adisyon Seçilmedi",
        font=("Segoe UI", 14, "bold")
    )
    baslik.pack(anchor="w", pady=(0, 10))

    tree = ttk.Treeview(
        sag,
        columns=("Ürün", "Adet", "Fiyat", "Tutar"),
        show="headings",
    )
    tree.tag_configure(
        "hazir",
        background="#d4f7d4",  # açık yeşil
        foreground="black"
    )

    for c in ("Ürün", "Adet", "Fiyat", "Tutar"):
        tree.heading(c, text=c)

    tree.column("Ürün", width=220)
    tree.column("Adet", width=80, anchor="center")
    tree.column("Fiyat", width=100, anchor="e")
    tree.column("Tutar", width=120, anchor="e")

    tree.pack(fill="both", expand=True, padx=5)

    # ================= SAĞ TIK MENÜSÜ (İPTAL) =================
    iptal_menu = tk.Menu(root, tearoff=0)
    iptal_menu.add_command(
        label="❌ Ürünü İptal Et",
        command=lambda: urun_iptal()
    )

    def siparis_sag_tik(event):
        secili = tree.identify_row(event.y)
        if secili:
            tree.selection_set(secili)
            iptal_menu.tk_popup(event.x_root, event.y_root)

    tree.bind("<Button-3>", siparis_sag_tik)

    def urun_iptal():
        secim = tree.selection()
        if not secim:
            return

        item = secim[0]
        urun_adi, adet, fiyat, tutar = tree.item(item)["values"]
        adet = int(adet)

        if not messagebox.askyesno(
            "Sipariş İptali",
            f"{urun_adi} (x{adet}) iptal edilsin mi?\n"
            "Stok geri eklenecek."
        ):
            return

        masa = baslik.cget("text").split("|")[0].replace(" Adisyonu", "").strip()

        # === MENU ID BUL ===
        menu_id = next(
            (k for k, v in menuler.items() if v["ad"] == urun_adi),
            None
        )
        if not menu_id:
            messagebox.showerror("Hata", "Menü bulunamadı")
            return

        # === BACKEND'E İPTAL BİLDİR ===
        siparis_iptal(masa, menu_id, adet, aktif_kullanici)

        # === STOK GERİ EKLE (REÇETEYE GÖRE) ===
        menu = menuler.get(menu_id)
        if menu:
            for r in menu.get("recete", []):
                urun = urunler[r["urun_kod"]]
                geri = r["miktar"] * adet

                urun["stok"] = round(urun["stok"] + geri, 2)

                hareketler.append({
                    "tarih": datetime.now().strftime("%d.%m.%Y %H:%M"),
                    "kullanici": aktif_kullanici,
                    "urun": urun["ad"],
                    "miktar": geri,
                    "tur": "GİRİŞ",
                    "aciklama": f"{urun_adi} iptal"
                })

            kaydet(DOSYA_URUN, urunler)
            kaydet(DOSYA_HAREKET, hareketler)

        # === TREEVIEW'DEN SİL ===
        tree.delete(item)

        # === TOPLAM GÜNCELLE ===
        toplam = sum(
            float(tree.item(i)["values"][3])
            for i in tree.get_children()
        )

        lbl_toplam_tutar.config(text=f"TOPLAM: {toplam:.2f} ₺")



    # ================= TOPLAM TUTAR =================
    alt_toplam = ttk.Frame(sag, padding=10)
    alt_toplam.pack(fill="x")

    lbl_toplam_tutar = ttk.Label(
        alt_toplam,
        text="TOPLAM: 0.00 ₺",
        font=("Segoe UI", 12, "bold")
    )
    lbl_toplam_tutar.pack(side="right")

    # ================= FONKSİYONLAR =================
    def masa_renk_guncelle(secili=None):
        nonlocal adisyonlar
        adisyonlar = adisyonlari_yukle()

        for masa, frame in masa_butonlari.items():
            ad = adisyonlar.get(masa)

            if masa == secili:
                renk = "#3498db"
            elif ad and ad.get("urunler"):
                renk = "#e74c3c"
            else:
                renk = "#2ecc71"

            frame.config(bg=renk)
            frame.winfo_children()[0].config(bg=renk)

    def indirim_uygula():
        global indirim_orani, indirimli_toplam

        try:
            oran = float(indirim_entry.get())
            if oran < 0 or oran > 100:
                raise ValueError
        except:
            messagebox.showerror("Hata", "İndirim yüzdesi 0-100 arasında olmalı")
            return

        # Ekrandaki TOPLAM'ı al
        try:
            mevcut_toplam = float(
                lbl_toplam_tutar.cget("text")
                .replace("TOPLAM:", "")
                .replace("₺", "")
                .strip()
            )
        except:
            return

        indirim_orani = oran
        indirimli_toplam = mevcut_toplam * (1 - oran / 100)

        lbl_toplam_tutar.config(
            text=f"TOPLAM: {indirimli_toplam:.2f} ₺  (İndirim %{oran})"
        )

    def masa_ac(masa):
        secili_masa["ad"] = masa

        tree.delete(*tree.get_children())

        toplam = 0.0
        adisyon = adisyonlar.get(masa)
        kullanici = adisyon.get("kullanici", "Bilinmiyor") if adisyon else "Bilinmiyor"

        baslik.config(text=f"{masa} Adisyonu | Garson: {kullanici}")

        if adisyon:
            for u in adisyon.get("urunler", []):
                menu = menuler.get(str(u["menu_id"]))
                if not menu:
                    continue

                t = u["adet"] * menu["fiyat"]
                toplam += t

                tags = ()
                if u.get("hazir") is True:
                    tags = ("hazir",)

                tree.insert(
                    "",
                    "end",
                    values=(
                        menu["ad"],
                        u["adet"],
                        f"{menu['fiyat']:.2f}",
                        f"{t:.2f}"
                    ),
                    tags=tags
                )

        lbl_toplam_tutar.config(text=f"TOPLAM: {toplam:.2f} ₺")
        masa_renk_guncelle(secili=masa)

    def menu_ekle():
        if baslik.cget("text") == "Adisyon Seçilmedi":
            messagebox.showwarning("Uyarı", "Önce masa seçin")
            return

        if not secili_menu_var.get():
            messagebox.showwarning("Uyarı", "Menü seçin")
            return

        try:
            a = int(adet.get())
        except:
            messagebox.showerror("Hata", "Adet geçersiz")
            return

        masa = baslik.cget("text").split("|")[0].replace(" Adisyonu", "").strip()
        mid = next(k for k, v in menuler.items() if v["ad"] == secili_menu_var.get())

        if not recete_stok_kontrol_ve_dus(mid, a):
            return

        siparis_gonder(masa, mid, a, aktif_kullanici)

        # 🔥 ADİSYONLARI TEKRAR YÜKLE
        adisyonlar.clear()
        adisyonlar.update(adisyonlari_yukle())

        # 🔥 ŞİMDİ GÜNCEL VERİYLE ÇİZ
        masa_ac(masa)

    def masa_kapat(tur):
        if baslik.cget("text") == "Adisyon Seçilmedi":
            return

        masa = baslik.cget("text").split("|")[0].replace(" Adisyonu", "").strip()

        backend_masa_kapat(masa, tur, aktif_kullanici)


        # ===== YAZDIRMA TÜRÜ SEÇ =====
        yazdirma_turu = yazici_sec_dialog()
        if not yazdirma_turu:
            return

        # ===== ÜRÜNLERİ TOPLA =====
        urun_listesi = []
        for item in tree.get_children():
            ad, adet, fiyat, tutar = tree.item(item)["values"]
            urun_listesi.append({
                "ad": ad,
                "adet": adet,
                "tutar": float(tutar)
            })

        toplam = float(
            lbl_toplam_tutar.cget("text")
            .split(":")[1]
            .replace("₺", "")
            .strip()
        )

        # ===== SADECE A4 (ŞİMDİLİK) =====
        if yazdirma_turu == "A4":
            pdf_yolu = pdf_fis_olustur(
                masa,
                urun_listesi,
                toplam,
                tur,
                aktif_kullanici
            )
            if not pdf_kaydet_dialog(pdf_yolu, masa):
                return

        # ===== EKRANI TEMİZLE =====
        tree.delete(*tree.get_children())
        baslik.config(text="Adisyon Seçilmedi")
        lbl_toplam_tutar.config(text="TOPLAM: 0.00 ₺")
        masa_renk_guncelle()


    # ================= ALT : SİPARİŞ =================
    alt = ttk.Frame(sag, padding=10)
    alt.pack(fill="x")

    secili_menu_var = tk.StringVar()
    adet = tk.StringVar(value="1")

    ttk.Label(alt, text="Menü").pack(side="left", padx=5)

    ttk.Combobox(
        alt,
        textvariable=secili_menu_var,
        values=[m["ad"] for m in menuler.values()],
        state="readonly",
        width=30
    ).pack(side="left", padx=5)

    ttk.Label(alt, text="Adet").pack(side="left", padx=5)
    ttk.Entry(alt, textvariable=adet, width=5).pack(side="left", padx=5)

    ttk.Button(alt, text="➕ Sipariş Ekle", command=menu_ekle).pack(side="left", padx=15)
    ttk.Button(alt, text="💳 Nakit ile Kapat", command=lambda: masa_kapat("Nakit")).pack(side="left", padx=5)
    ttk.Button(alt, text="💳 Kart ile Kapat", command=lambda: masa_kapat("Kart")).pack(side="left", padx=5)

    # ================= İNDİRİM =================
    ttk.Label(alt, text="İndirim %").pack(side="left", padx=(20, 5))

    indirim_entry = ttk.Entry(alt, width=5)
    indirim_entry.pack(side="left")

    ttk.Button(
        alt,
        text="İndirim Uygula",
        command=indirim_uygula
    ).pack(side="left", padx=5)


    # ================= MASA BUTONLARI =================
    row, col = 1, 0
    for masa in masalar:
        f = tk.Frame(sol, bg="#2ecc71", padx=2, pady=2)
        f.grid(row=row, column=col, sticky="ew", padx=2, pady=2)

        btn = tk.Button(
            f,
            text=masa,
            font=("Segoe UI", 9, "bold"),
            relief="flat",
            command=lambda m=masa: masa_ac(m)
        )
        btn.pack(fill="x")

        def sag_tik(event, m=masa):
            if m in adisyonlar:  # sadece dolu masa taşınır
                masa_menu.delete(0, "end")
                masa_menu.add_command(
                    label="🔀 Masayı Taşı",
                    command=lambda: masa_tasi_dialog(m)
                )
                masa_menu.tk_popup(event.x_root, event.y_root)

        btn.bind("<Button-3>", sag_tik)


        masa_butonlari[masa] = f

        col += 1
        if col >= 2:
            col = 0
            row += 1

    masa_renk_guncelle()

    def oto_adisyon_guncelle():
        try:
            yeni = adisyonlari_yukle()

            if yeni != adisyonlar:
                adisyonlar.clear()
                adisyonlar.update(yeni)

                masa_renk_guncelle()

                if secili_masa["ad"]:
                    masa_ac(secili_masa["ad"])

        except Exception as e:
            print("OTO GÜNCELLE HATA:", e)

        root.after(3000, oto_adisyon_guncelle)


    oto_adisyon_guncelle()

def kullanici_yonetimi():
    if aktif_rol != "admin":
        messagebox.showerror("Yetki", "Sadece admin erişebilir")
        return

    temizle_orta_alan()

    frame = ttk.Frame(content_frame, padding=20)
    frame.pack(fill="both", expand=True)

    ttk.Label(
        frame,
        text="👥 Kullanıcı Yönetimi",
        font=("Segoe UI", 16, "bold")
    ).pack(pady=10)

    # ================= TABLO =================
    tree = ttk.Treeview(
        frame,
        columns=("Kullanıcı", "Rol", "Durum"),
        show="headings",
        height=12
    )
    tree.heading("Kullanıcı", text="Kullanıcı")
    tree.heading("Rol", text="Rol")
    tree.heading("Durum", text="Durum")

    tree.column("Kullanıcı", width=220, anchor="w")
    tree.column("Rol", width=120, anchor="center")
    tree.column("Durum", width=120, anchor="center")
    tree.pack(fill="both", expand=True, pady=10)

    def yenile():
        tree.delete(*tree.get_children())
        for ad, u in kullanicilar.items():
            durum = "Aktif" if u.get("aktif", True) else "Pasif"
            tree.insert("", "end", iid=ad, values=(ad, u["role"], durum))

    yenile()

    # ================= KULLANICI EKLE =================
    def kullanici_ekle():
        p = tk.Toplevel(root)
        p.title("Kullanıcı Ekle")
        p.geometry("480x650")
        p.grab_set()

        ad = tk.StringVar()
        sifre = tk.StringVar()
        rol = tk.StringVar(value="garson")

        ttk.Label(p, text="Kullanıcı Adı").pack(pady=4)
        ttk.Entry(p, textvariable=ad).pack(fill="x", padx=20)

        ttk.Label(p, text="Şifre").pack(pady=4)
        ttk.Entry(p, textvariable=sifre, show="*").pack(fill="x", padx=20)

        ttk.Label(p, text="Rol").pack(pady=4)
        ttk.Combobox(
            p,
            textvariable=rol,
            values=["admin", "garson"],
            state="readonly"
        ).pack(fill="x", padx=20)

        ttk.Label(p, text="Yetkiler", font=FONT_BIG).pack(anchor="w", padx=20, pady=(10, 5))

        yetki_tree = ttk.Treeview(p, show="tree")
        yetki_tree.pack(fill="both", expand=True, padx=20, pady=5)

        # === MENÜLER ===
        for ana, altlar in MENU_YAPISI.items():
            parent = yetki_tree.insert("", "end", text=f"☐ {ana}", open=True)
            for alt in altlar.keys():
                yetki_tree.insert(parent, "end", text=f"☐ {alt}")

        # === CHECKBOX TOGGLE ===
        def toggle(event):
            item = yetki_tree.identify_row(event.y)
            if not item:
                return

            text = yetki_tree.item(item, "text")
            if not text.startswith(("☐", "☑")):
                return

            secili = text.startswith("☑")
            yeni_text = ("☐ " if secili else "☑ ") + text[2:]
            yetki_tree.item(item, text=yeni_text)

            # ana menüyse altları da değiştir
            for child in yetki_tree.get_children(item):
                ctext = yetki_tree.item(child, "text")
                yetki_tree.item(
                    child,
                    text=("☐ " if secili else "☑ ") + ctext[2:]
                )

        yetki_tree.bind("<ButtonRelease-1>", toggle)

        # === KAYDET ===
        def kaydet_kullanici():
            if not ad.get() or not sifre.get():
                messagebox.showerror("Hata", "Alanlar boş")
                return

            if any(k.lower() == ad.get().lower() for k in kullanicilar):
                messagebox.showerror("Hata", "Kullanıcı mevcut")
                return

            yetkiler = {}
            for pitem in yetki_tree.get_children():
                ana = yetki_tree.item(pitem, "text")[2:]
                yetkiler[ana] = {}
                for c in yetki_tree.get_children(pitem):
                    alt = yetki_tree.item(c, "text")[2:]
                    yetkiler[ana][alt] = yetki_tree.item(c, "text").startswith("☑")

            # ADMIN ise tüm yetkileri otomatik aç
            if rol.get() == "admin":
                for ana in yetkiler:
                    for alt in yetkiler[ana]:
                        yetkiler[ana][alt] = True

            kullanicilar[ad.get()] = {
                "password": hashle(sifre.get()),
                "role": rol.get(),
                "aktif": True,
                "yetkiler": yetkiler
            }

            kaydet(DOSYA_USER, kullanicilar)
            yenile()
            p.destroy()

        ttk.Button(p, text="💾 Kaydet", command=kaydet_kullanici).pack(pady=15)


    def kullanici_duzenle():
        sec = tree.focus()
        if not sec:
            messagebox.showwarning("Uyarı", "Düzenlenecek kullanıcıyı seçin")
            return

        veri = kullanicilar[sec]

        p = tk.Toplevel(root)
        p.title("Kullanıcı Düzenle")
        p.geometry("480x650")
        p.grab_set()

        ad = tk.StringVar(value=sec)
        sifre = tk.StringVar()
        rol = tk.StringVar(value=veri["role"])
        aktif_var = tk.BooleanVar(value=veri.get("aktif", True))

        ttk.Label(p, text="Kullanıcı Adı").pack(pady=4)
        ttk.Entry(p, textvariable=ad, state="disabled").pack(fill="x", padx=20)

        ttk.Label(p, text="Yeni Şifre (boş bırakılırsa değişmez)")\
            .pack(pady=4)
        ttk.Entry(p, textvariable=sifre, show="*")\
            .pack(fill="x", padx=20)

        ttk.Label(p, text="Rol").pack(pady=4)
        ttk.Combobox(
            p,
            textvariable=rol,
            values=["admin", "garson"],
            state="readonly"
        ).pack(fill="x", padx=20)
        ttk.Checkbutton(
            p,
            text="Kullanıcı Aktif",
            variable=aktif_var
        ).pack(anchor="w", padx=20, pady=5)

        ttk.Label(p, text="Yetkiler", font=FONT_BIG)\
            .pack(anchor="w", padx=20, pady=(10, 5))

        yetki_tree = ttk.Treeview(p, show="tree")
        yetki_tree.pack(fill="both", expand=True, padx=20, pady=5)

        # === YETKİLERİ DOLDUR ===
        for ana, altlar in MENU_YAPISI.items():
            parent = yetki_tree.insert("", "end", text=f"☐ {ana}", open=True)

            for alt in altlar.keys():
                aktif = veri["yetkiler"].get(ana, {}).get(alt, False)
                ikon = "☑" if aktif else "☐"
                yetki_tree.insert(parent, "end", text=f"{ikon} {alt}")

        # === TOGGLE ===
        def toggle(event):
            item = yetki_tree.identify_row(event.y)
            if not item:
                return

            text = yetki_tree.item(item, "text")
            if not text.startswith(("☐", "☑")):
                return

            secili = text.startswith("☑")
            yeni = ("☐ " if secili else "☑ ") + text[2:]
            yetki_tree.item(item, text=yeni)

            for c in yetki_tree.get_children(item):
                ct = yetki_tree.item(c, "text")
                yetki_tree.item(c, text=("☐ " if secili else "☑ ") + ct[2:])

        yetki_tree.bind("<ButtonRelease-1>", toggle)

        # === KAYDET ===
        def kaydet_degisim():
            yetkiler = {}

            for pitem in yetki_tree.get_children():
                ana = yetki_tree.item(pitem, "text")[2:]
                yetkiler[ana] = {}

                for c in yetki_tree.get_children(pitem):
                    alt = yetki_tree.item(c, "text")[2:]
                    yetkiler[ana][alt] = yetki_tree.item(c, "text").startswith("☑")

            veri["role"] = rol.get()
            veri["yetkiler"] = yetkiler
            veri["aktif"] = aktif_var.get()

            if sifre.get().strip():
                veri["password"] = hashle(sifre.get())

            kaydet(DOSYA_USER, kullanicilar)
            yenile()
            p.destroy()

        ttk.Button(p, text="💾 Kaydet", command=kaydet_degisim).pack(pady=15)
    
    def kullanici_sil():
        sec = tree.focus()
        if not sec:
            messagebox.showwarning("Uyarı", "Silinecek kullanıcıyı seçin")
            return

        if sec.lower() == "admin":
            messagebox.showerror("Hata", "Admin silinemez")
            return

        if not messagebox.askyesno(
            "Kullanıcı Sil",
            f"{sec} kullanıcısı kalıcı olarak silinsin mi?"
        ):
            return

        kullanicilar.pop(sec)
        kaydet(DOSYA_USER, kullanicilar)
        yenile()


    # ================= ALT BUTONLAR =================
    btn = ttk.Frame(frame)
    btn.pack(pady=15)

    ttk.Button(btn, text="➕ Kullanıcı Ekle", width=18, command=kullanici_ekle)\
        .pack(side="left", padx=6)

    ttk.Button(btn, text="✏️ Kullanıcı Düzenle", width=18, command=kullanici_duzenle)\
        .pack(side="left", padx=6)

    ttk.Button(
        btn,
        text="🗑 Kullanıcı Sil",
        width=18,
        command=kullanici_sil
    ).pack(side="left", padx=6)




def dashboard():
    
    global gelirler, giderler, urunler

    temizle_orta_alan()

    frame = tk.Frame(content_frame, bg="white")
    frame.pack(fill="both", expand=True, padx=40, pady=40)

    today = datetime.now().strftime("%d.%m.%Y")

    gunluk_gelir = sum(g["tutar"] for g in gelirler if g["tarih"].startswith(today))

    gunluk_gider = sum(g["tutar"] for g in giderler if g["tarih"].startswith(today))

    net = gunluk_gelir - gunluk_gider

    kritik_sayisi = sum(1 for u in urunler.values() if u["stok"] <= u.get("kritik", 0))

    frame.columnconfigure((0, 1), weight=1)
    frame.rowconfigure((0, 1), weight=1)

    kart_font = ("Segoe UI", 22, "bold")
    baslik_font = ("Segoe UI", 14)

    def kart(row, col, baslik, deger, renk):
        f = tk.Frame(frame, bg=renk, bd=0)
        f.grid(row=row, column=col, padx=20, pady=20, sticky="nsew")

        tk.Label(f, text=baslik, bg=renk, fg="white", font=baslik_font).pack(
            pady=(30, 10)
        )

        tk.Label(f, text=deger, bg=renk, fg="white", font=kart_font).pack(pady=(0, 30))

    kart(0, 0, "Bugünkü Gelir", f"{gunluk_gelir:.2f} ₺", "#2ecc71")
    kart(0, 1, "Bugünkü Gider", f"{gunluk_gider:.2f} ₺", "#e74c3c")
    kart(1, 0, "Net Sonuç", f"{net:.2f} ₺", "#3498db")
    kart(1, 1, "Kritik Stok", f"{kritik_sayisi} Ürün", "#f39c12")


def ana_sayfa_goster():
    temizle_orta_alan()

    try:
        logo_path = os.path.join(BASE_DIR, "ceviz.png")
        img = Image.open(logo_path)
        img = img.resize((500, 500), Image.LANCZOS)
        logo_img = ImageTk.PhotoImage(img)

        lbl = tk.Label(content_frame, image=logo_img, bg="white")
        lbl.image = logo_img
        lbl.pack(expand=True)

    except Exception as e:
        ttk.Label(
            content_frame, text="Ceviz Altı Restaurant", font=("Segoe UI", 24, "bold")
        ).pack(expand=True)


# ================= ANA EKRAN =================
def ana_ekran():
    aktif_yetkiler = kullanicilar.get(aktif_kullanici, {}).get("yetkiler", {})
    global urunler, hareketler, giderler, gelirler
    global menuler

    menuler = yukle(DOSYA_MENU, {})

    root.title(f"Ceviz Altı Restaurant - {aktif_kullanici} ({aktif_rol})")

    urunler = yukle(DOSYA_URUN, {})
    hareketler = yukle(DOSYA_HAREKET, [])
    giderler = yukle(DOSYA_GIDER, [])
    gelirler = yukle(DOSYA_GELIR, [])
    satis_detay = yukle(DOSYA_SATIS_DETAY, [])


    ana_sayfa_goster()
    
    def sag_tik_satin_al(secili_urun):
        p = tk.Toplevel(root)
        p.title("🛒 Satın Al / Gider Girişi")
        p.geometry("420x360")
        p.grab_set()

        miktar = tk.StringVar(value="1")
        birim_fiyat = tk.StringVar(value="0")
        toplam = tk.StringVar(value="0.00")

        def hesapla(*_):
            try:
                toplam.set(f"{float(miktar.get()) * float(birim_fiyat.get()):.2f}")
            except:
                toplam.set("0.00")

        miktar.trace_add("write", hesapla)
        birim_fiyat.trace_add("write", hesapla)

        ttk.Label(p, text=f"Ürün: {secili_urun}", font=FONT_BIG).pack(pady=10)

        ttk.Label(p, text="Miktar").pack(anchor="w", padx=30)
        ttk.Entry(p, textvariable=miktar).pack(fill="x", padx=30)

        ttk.Label(p, text="Birim Fiyat (₺)").pack(anchor="w", padx=30, pady=(10, 0))
        ttk.Entry(p, textvariable=birim_fiyat).pack(fill="x", padx=30)

        ttk.Label(p, text="Toplam (₺)").pack(anchor="w", padx=30, pady=(10, 0))
        ttk.Entry(p, textvariable=toplam, state="readonly").pack(fill="x", padx=30)

        def kaydet_satin_al_pencere():
            try:
                m = float(miktar.get())
                f = float(birim_fiyat.get())
            except:
                messagebox.showerror("Hata", "Miktar veya fiyat geçersiz")
                return

            kod = next(k for k, v in urunler.items() if v["ad"] == secili_urun)

            # ✅ STOK ARTIR
            urunler[kod]["stok"] = round(urunler[kod]["stok"] + m, 2)
            urunler[kod]["son_alis_fiyat"] = f

            # ✅ GİDER YAZ
            giderler.append({
                "tarih": datetime.now().strftime("%d.%m.%Y %H:%M"),
                "kategori": "Malzeme",
                "urun": secili_urun,
                "aciklama": "Satın alma",
                "miktar": m,
                "tutar": m * f,
                "kullanici": aktif_kullanici
            })

            # ✅ HAREKET (GİRİŞ)
            hareketler.append({
                "tarih": datetime.now().strftime("%d.%m.%Y %H:%M"),
                "kullanici": aktif_kullanici,
                "urun": secili_urun,
                "miktar": m,
                "tur": "GİRİŞ"
            })

            kaydet(DOSYA_URUN, urunler)
            kaydet(DOSYA_GIDER, giderler)
            kaydet(DOSYA_HAREKET, hareketler)

            messagebox.showinfo("Başarılı", "Satın alma kaydedildi")
            p.destroy()
            stok_goster()

        ttk.Button(p, text="💾 Kaydet", command=kaydet_satin_al_pencere).pack(pady=25)

    # ---------- DEPO ----------
    def stok_goster():
        temizle_orta_alan()

        frame = ttk.Frame(content_frame, padding=20)
        frame.pack(fill="both", expand=True)


        def sirala(tree, col, numeric):
            data = [(tree.set(k, col), k) for k in tree.get_children("")]
            if numeric:
                data.sort(key=lambda t: float(t[0]))
            else:
                data.sort(key=lambda t: t[0].lower())
            for index, (_, k) in enumerate(data):
                tree.move(k, "", index)

        tree = ttk.Treeview(
            frame,
            columns=("Ürün", "Stok", "Birim", "Kritik"),
            show="headings"
        )
        tree.pack(fill="both", expand=True)

        tree.heading("Ürün", text="Ürün Adı", command=lambda: sirala(tree, "Ürün", False))
        tree.heading("Stok", text="Stok", command=lambda: sirala(tree, "Stok", True))
        tree.heading("Birim", text="Birim", command=lambda: sirala(tree, "Birim", False))
        tree.heading("Kritik", text="Kritik", command=lambda: sirala(tree, "Kritik", True))

        tree.column("Ürün", width=300, anchor="w")
        tree.column("Stok", width=130, anchor="center")
        tree.column("Birim", width=120, anchor="center")
        tree.column("Kritik", width=120, anchor="center")

        tree.tag_configure("kritik", background="#ffd6d6")

        stok_menu = tk.Menu(root, tearoff=0)
        stok_menu.add_command(
            label="➕ Satın Al (Gider Gir)",
            command=lambda: sag_tik_satin_al(
                tree.item(tree.selection()[0])["values"][0]
            )
        )

        def sag_tik_menu(event):
            secili = tree.identify_row(event.y)
            if secili:
                tree.selection_set(secili)
                stok_menu.tk_popup(event.x_root, event.y_root)

        tree.bind("<Button-3>", sag_tik_menu)

        for u in urunler.values():
            tag = ()
            if u["stok"] <= u.get("kritik", 0):
                tag = ("kritik",)

            tree.insert(
                "",
                "end",
                values=(
                    u["ad"],
                    f"{u['stok']:.2f}",
                    u["birim"],
                    u.get("kritik", 0),
                ),
                tags=tag,
            )


    def stok_hareket(tur, secili_urun=None):
        FONT_LABEL = ("Segoe UI", 20, "bold")
        FONT_ENTRY = ("Segoe UI", 20)
        FONT_TOPLAM = ("Segoe UI", 23, "bold")

        temizle_orta_alan()

        main_frame = ttk.Frame(content_frame)
        main_frame.pack(fill="both", expand=True)

        urun = tk.StringVar()
        miktar = tk.StringVar()
        birim = tk.StringVar()
        birim_fiyat = tk.StringVar()

        form_frame = ttk.LabelFrame(
            main_frame,
            text="Stok Çıkışı",
            padding=(40, 35)
        )
        form_frame.place(relx=0.5, rely=0.5, anchor="center")

        # ÜRÜN
        ttk.Label(form_frame, text="Ürün", font=FONT_LABEL)\
            .grid(row=0, column=0, sticky="w", pady=(0, 6))

        urun_cb = ttk.Combobox(
            form_frame,
            textvariable=urun,
            values=[u["ad"] for u in urunler.values()],
            state="readonly",
            font=FONT_ENTRY,
            width=40
        )
        urun_cb.grid(row=1, column=0, pady=(0, 14))

        birim_lbl = ttk.Label(
            form_frame,
            text="Birim: -",
            font=("Segoe UI", 12, "italic"),
            foreground="#555"
        )
        birim_lbl.grid(row=2, column=0, sticky="w", pady=(0, 18))

        def urun_degisti(event=None):
            for u in urunler.values():
                if u["ad"] == urun.get():
                    birim_lbl.config(text=f"Birim: {u.get('birim','')}")
                    return

        urun_cb.bind("<<ComboboxSelected>>", urun_degisti)

        if secili_urun:
            urun.set(secili_urun)
            form_frame.after(50, urun_degisti)

        # MİKTAR
        ttk.Label(form_frame, text="Miktar", font=FONT_LABEL)\
            .grid(row=3, column=0, sticky="w", pady=(0, 6))

        ttk.Entry(
            form_frame,
            textvariable=miktar,
            font=FONT_ENTRY,
            width=32
        ).grid(row=4, column=0, pady=(0, 16))

        # BİRİM FİYAT (SADECE GİRİŞTE)
        if tur == "GİRİŞ":
            ttk.Label(
                form_frame,
                text="Birim Fiyat (₺)",
                font=FONT_LABEL
            ).grid(row=5, column=0, sticky="w", pady=(0, 6))

            ttk.Entry(
                form_frame,
                textvariable=birim_fiyat,
                font=FONT_ENTRY,
                width=32
            ).grid(row=6, column=0, pady=(0, 18))

        ttk.Button(
            form_frame,
            text="💾 Kaydet",
            width=22,
            command=lambda: kaydet_satin_al()
        ).grid(row=8, column=0)

        def kaydet_satin_al():
            # === MİKTAR KONTROL ===
            try:
                m = float(miktar.get())
                if m <= 0:
                    raise ValueError
            except:
                messagebox.showerror("Hata", "Miktar geçersiz")
                return

            # === ÜRÜN ADI ===
            urun_adi = urun.get()
            if not urun_adi:
                messagebox.showerror("Hata", "Ürün seçilmedi")
                return

            # === ÜRÜN KODU ===
            kod = next(
                (k for k, v in urunler.items() if v["ad"] == urun_adi),
                None
            )
            if not kod:
                messagebox.showerror("Hata", "Ürün kodu bulunamadı")
                return

            # === FİYAT BELİRLEME ===
            if tur == "ÇIKIŞ":  # 🔴 ÇÖP / FİRE
                f = urunler[kod].get("son_alis_fiyat", 0)
                if f <= 0:
                    messagebox.showerror(
                        "Hata",
                        "Bu ürün için son alış fiyatı yok.\nÖnce satın alma yapmalısınız."
                    )
                    return
            else:  # 🔵 GİRİŞ
                try:
                    f = float(birim_fiyat.get())
                    if f <= 0:
                        raise ValueError
                except:
                    messagebox.showerror("Hata", "Birim fiyat geçersiz")
                    return

            toplam_tutar = round(m * f, 2)

            # === STOK GÜNCELLE ===
            if tur == "GİRİŞ":
                urunler[kod]["stok"] = round(urunler[kod]["stok"] + m, 2)
                urunler[kod]["son_alis_fiyat"] = f
            else:  # ÇIKIŞ
                if urunler[kod]["stok"] < m:
                    messagebox.showerror("Hata", "Yetersiz stok")
                    return
                urunler[kod]["stok"] = round(urunler[kod]["stok"] - m, 2)

            # === GİDER KAYDI ===
            giderler.append({
                "tarih": datetime.now().strftime("%d.%m.%Y %H:%M"),
                "kategori": "Malzeme",
                "urun": urun_adi,
                "aciklama": "Satın alma" if tur == "GİRİŞ" else "Fire / Çöp",
                "miktar": m,
                "tutar": toplam_tutar,
                "kullanici": aktif_kullanici
            })

            # === STOK HAREKETİ ===
            hareketler.append({
                "tarih": datetime.now().strftime("%d.%m.%Y %H:%M"),
                "kullanici": aktif_kullanici,
                "urun": urun_adi,
                "miktar": m,
                "tur": tur
            })

            # === DOSYALARA KAYDET ===
            kaydet(DOSYA_URUN, urunler)
            kaydet(DOSYA_GIDER, giderler)
            kaydet(DOSYA_HAREKET, hareketler)

            # === BİLGİ ===
            messagebox.showinfo(
                "Başarılı",
                f"{urun_adi}\n\n"
                f"İşlem: {tur}\n"
                f"Miktar: {m}\n"
                f"Birim Fiyat: {f:.2f} ₺\n"
                f"Toplam: {toplam_tutar:.2f} ₺"
            )

            stok_goster()

            messagebox.showinfo("Başarılı", "Stok güncellendi")
            stok_goster()

    def stok_hareketleri():
        temizle_orta_alan()

        main_frame = ttk.Frame(content_frame, padding=20)
        main_frame.pack(fill="both", expand=True)

        # ================= FİLTRE BAR =================
        filter_frame = ttk.Frame(main_frame)
        filter_frame.pack(fill="x", pady=(0, 15))

        baslangic_var = tk.StringVar()
        bitis_var = tk.StringVar()

        ttk.Label(filter_frame, text="Başlangıç").pack(side="left", padx=5)
        DateEntry(filter_frame, textvariable=baslangic_var,
                date_pattern="dd.MM.yyyy", width=12).pack(side="left")

        ttk.Label(filter_frame, text="Bitiş").pack(side="left", padx=5)
        DateEntry(filter_frame, textvariable=bitis_var,
                date_pattern="dd.MM.yyyy", width=12).pack(side="left")

        ttk.Button(filter_frame, text="🔍 Filtrele",
                command=lambda: hareketleri_yukle()).pack(side="left", padx=15)

        # ================= TABLOLAR (GRID) =================
        tables = ttk.Frame(main_frame)
        tables.pack(fill="both", expand=True)

        tables.columnconfigure(0, weight=1, uniform="x")
        tables.columnconfigure(1, weight=1, uniform="x")
        tables.rowconfigure(0, weight=1)

        KOLONLAR = ("Tarih", "Kullanıcı", "Ürün", "Miktar")
        GENISLIK = {
            "Tarih": 160,
            "Kullanıcı": 100,
            "Ürün": 220,
            "Miktar": 80
        }

        # ================= GİRİŞLER =================
        giris_frame = ttk.LabelFrame(tables, text="GİRİŞLER", padding=10)
        giris_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5))

        giris_tree = ttk.Treeview(giris_frame, columns=KOLONLAR, show="headings")
        giris_tree.pack(fill="both", expand=True)

        for c in KOLONLAR:
            giris_tree.heading(c, text=c)
            giris_tree.column(
                c,
                width=GENISLIK[c],
                anchor="w" if c == "Ürün" else "center",
                stretch=False
            )

        giris_tree.column("Miktar", anchor="e")
        giris_tree.tag_configure("giris", background="#e6ffe6")

        # ================= ÇIKIŞLAR =================
        cikis_frame = ttk.LabelFrame(tables, text="ÇIKIŞLAR", padding=10)
        cikis_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0))

        cikis_tree = ttk.Treeview(cikis_frame, columns=KOLONLAR, show="headings")
        cikis_tree.pack(fill="both", expand=True)

        for c in KOLONLAR:
            cikis_tree.heading(c, text=c)
            cikis_tree.column(
                c,
                width=GENISLIK[c],
                anchor="w" if c == "Ürün" else "center",
                stretch=False
            )

        cikis_tree.column("Miktar", anchor="e")
        cikis_tree.tag_configure("cikis", background="#ffe6e6")

        # ================= VERİ YÜKLE =================
        def hareketleri_yukle():
            giris_tree.delete(*giris_tree.get_children())
            cikis_tree.delete(*cikis_tree.get_children())

            try:
                b1 = datetime.strptime(baslangic_var.get(), "%d.%m.%Y").date()
                b2 = datetime.strptime(bitis_var.get(), "%d.%m.%Y").date()
            except:
                messagebox.showerror("Hata", "Tarih seçiniz")
                return

            for h in hareketler:
                try:
                    ht = datetime.strptime(h["tarih"], "%d.%m.%Y %H:%M").date()
                except:
                    continue

                if not (b1 <= ht <= b2):
                    continue

                row = (
                    h["tarih"],
                    h["kullanici"],
                    h["urun"],
                    f'{h["miktar"]:+}'
                )

                if h["tur"] == "GİRİŞ":
                    giris_tree.insert("", "end", values=row, tags=("giris",))
                else:
                    cikis_tree.insert("", "end", values=row, tags=("cikis",))

        bugun = datetime.now().strftime("%d.%m.%Y")
        baslangic_var.set(bugun)
        bitis_var.set(bugun)
        hareketleri_yukle()


    def urun_yonetimi():
        if aktif_rol != "admin":
            messagebox.showerror("Yetki", "Sadece admin")
            return

        temizle_orta_alan()

        # ANA FRAME
        main_frame = ttk.Frame(content_frame, padding=20)
        main_frame.pack(fill="both", expand=True)

        # TABLO
        tree = ttk.Treeview(
            main_frame,
            columns=("Kod", "Ad","Tip", "Birim", "Stok", "Kritik"),
            show="headings",
        )

        for c in ("Kod", "Ad", "Tip", "Birim", "Stok", "Kritik"):
            tree.heading(c, text=c)

        tree.column("Kod", width=80)
        tree.column("Ad", width=200)
        tree.column("Tip", width=100)
        tree.column("Birim", width=120)
        tree.column("Stok", width=100, anchor="e")
        tree.column("Kritik", width=100, anchor="e")

        tree.pack(fill="both", expand=True, pady=10)

        # KRİTİK RENK
        tree.tag_configure("kritik", background="#ffd6d6")

        def yenile():
            tree.delete(*tree.get_children())

            for k in sorted(urunler.keys(), key=lambda x: int(x)):
                v = urunler[k]

                tag = ()
                if v["stok"] <= v.get("kritik", 0):
                    tag = ("kritik",)

                tree.insert(
                    "",
                    "end",
                    values=(
                        k,
                        v["ad"],
                        v.get("tip", "❌ YOK"),
                        v["birim"],
                        v["stok"],
                        v.get("kritik", 0),
                    ),
                    tags=tag,
                )


        # ---------- ÜRÜN EKLE ----------
        def urun_ekle():
            p2 = tk.Toplevel(root)
            p2.title("Ürün Ekle")
            p2.geometry("400x360")
            p2.focus_force()
            p2.resizable(False, False)
            p2.transient(root)
            p2.grab_set()

            frame = ttk.Frame(p2, padding=20)
            frame.pack(fill="both", expand=True)

            ad = tk.StringVar()
            birim = tk.StringVar()
            kritik = tk.StringVar()
            tip = tk.StringVar(value="yiyecek")

            ttk.Label(frame, text="Ürün Adı").grid(row=0, column=0, sticky="w", pady=5)
            ttk.Entry(frame, textvariable=ad, width=30).grid(row=0, column=1, pady=5)

            ttk.Label(frame, text="Birim").grid(row=1, column=0, sticky="w", pady=5)
            ttk.Combobox(
                frame,
                textvariable=birim,
                values=["Adet", "Kg", "Litre"],
                state="readonly",
                width=28
            ).grid(row=1, column=1, pady=5)

            ttk.Label(frame, text="Ürün Tipi").grid(row=2, column=0, sticky="w", pady=5)
            ttk.Combobox(
                frame,
                textvariable=tip,
                values=["Yiyecek", "İçecek", "Diğer"],
                state="readonly",
                width=28
            ).grid(row=2, column=1, pady=5)

            ttk.Label(frame, text="Kritik Seviye").grid(row=3, column=0, sticky="w", pady=5)
            ttk.Entry(frame, textvariable=kritik, width=30).grid(row=3, column=1, pady=5)

            def kaydet_urun():
                if not ad.get() or not birim.get():
                    messagebox.showerror("Hata", "Ürün adı ve birim zorunlu")
                    return

                try:
                    kritik_deger = int(kritik.get()) if kritik.get() else 0
                except:
                    messagebox.showerror("Hata", "Kritik seviye sayı olmalı")
                    return

                kod = str(max(map(int, urunler.keys()), default=0) + 1)

                urunler[kod] = {
                    "ad": ad.get().strip(),
                    "birim": birim.get().strip(),
                    "stok": 0,
                    "kritik": kritik_deger,
                    "son_alis_fiyat": 0,
                    "tip": tip.get().lower().replace("ı", "i")
                }

                kaydet(DOSYA_URUN, urunler)
                yenile()
                p2.destroy()

            btn_frame = ttk.Frame(frame)
            btn_frame.grid(row=4, column=0, columnspan=2, pady=20)

            ttk.Button(btn_frame, text="💾 Kaydet", command=kaydet_urun)\
                .pack(side="left", padx=5)
            ttk.Button(btn_frame, text="İptal", command=p2.destroy)\
                .pack(side="left", padx=5)

            

        # ---------- ÜRÜN SİL ----------
        def urun_sil():
            sec = tree.focus()
            if not sec:
                messagebox.showwarning("Uyarı", "Silinecek ürünü seçin")
                return

            kod = tree.item(sec)["values"][0]
            ad = tree.item(sec)["values"][1]

            if not messagebox.askyesno("Ürün Sil", f"'{ad}' ürünü silinsin mi?"):
                return

            urunler.pop(str(kod), None)
            kaydet(DOSYA_URUN, urunler)
            yenile()

        def urun_duzenle():
            secim = tree.selection()
            if not secim:
                messagebox.showwarning("Uyarı", "Düzenlenecek ürünü seçin")
                return

            sec = secim[0]
            kod, ad_eski, tip_eski, birim_eski, stok, kritik_eski = tree.item(sec)["values"]

            p2 = tk.Toplevel(root)
            p2.title("Ürün Düzenle")
            p2.geometry("420x420")
            p2.transient(root)
            p2.grab_set()

            frame = ttk.Frame(p2, padding=20)
            frame.pack(fill="both", expand=True)

            kod_var = tk.StringVar(value=kod)
            ad = tk.StringVar(value=ad_eski)
            tip = tk.StringVar(value=tip_eski)
            birim = tk.StringVar(value=birim_eski)
            kritik = tk.StringVar(value=kritik_eski)

            ttk.Label(frame, text="Ürün Kodu").grid(row=0, column=0, sticky="w", pady=5)
            ttk.Entry(frame, textvariable=kod_var, width=30).grid(row=0, column=1, pady=5)

            ttk.Label(frame, text="Ürün Adı").grid(row=1, column=0, sticky="w", pady=5)
            ttk.Entry(frame, textvariable=ad, width=30).grid(row=1, column=1, pady=5)

            ttk.Label(frame, text="Birim").grid(row=2, column=0, sticky="w", pady=5)
            ttk.Combobox(
                frame,
                textvariable=birim,
                values=["Adet", "Kg", "Litre"],
                state="readonly",
                width=28
            ).grid(row=2, column=1, pady=5)

            ttk.Label(frame, text="Ürün Tipi").grid(row=3, column=0, sticky="w", pady=5)
            ttk.Combobox(
                frame,
                textvariable=tip,
                values=["Yiyecek", "İçecek", "Diğer"],
                state="readonly",
                width=28
            ).grid(row=3, column=1, pady=5)

            ttk.Label(frame, text="Kritik Seviye").grid(row=4, column=0, sticky="w", pady=5)
            ttk.Entry(frame, textvariable=kritik, width=30).grid(row=4, column=1, pady=5)

            btn_frame2 = ttk.Frame(frame)
            btn_frame2.grid(row=5, column=0, columnspan=2, pady=25)

            def kaydet_duzenleme():
                yeni_kod = kod_var.get().strip()

                if not yeni_kod:
                    messagebox.showerror("Hata", "Kod boş olamaz")
                    return

                if yeni_kod != str(kod) and yeni_kod in urunler:
                    messagebox.showerror("Hata", "Bu kod zaten kullanılıyor")
                    return

                try:
                    kritik_deger = int(kritik.get())
                except:
                    messagebox.showerror("Hata", "Kritik seviye sayı olmalı")
                    return

                if yeni_kod != str(kod):
                    urunler[yeni_kod] = urunler.pop(str(kod))

                urunler[yeni_kod]["ad"] = ad.get().strip()
                urunler[yeni_kod]["tip"] = tip.get().lower().replace("ı", "i")
                urunler[yeni_kod]["birim"] = birim.get()
                urunler[yeni_kod]["kritik"] = kritik_deger
                

                kaydet(DOSYA_URUN, urunler)
                yenile()
                p2.destroy()

            ttk.Button(btn_frame2, text="💾 Kaydet", command=kaydet_duzenleme).pack(
                side="left", padx=10
            )
            ttk.Button(btn_frame2, text="İptal", command=p2.destroy).pack(
                side="left", padx=10
            )


        # ---------- BUTONLAR ----------
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=10)

        ttk.Button(btn_frame, text="➕ Ürün Ekle", command=urun_ekle).pack(
            side="left", padx=10
        )
        ttk.Button(btn_frame, text="🗑 Ürün Sil", command=urun_sil).pack(
            side="left", padx=10
        )
        ttk.Button(btn_frame, text="✏️ Ürün Düzenle", command=urun_duzenle).pack(
            side="left", padx=10)
        
        yenile()



    # ---------- SATIN ALMA ----------
    def gider_giris_ekrani():
        temizle_orta_alan()

        ana = ttk.Frame(content_frame, padding=20)
        ana.pack(fill="both", expand=True)

        # ================= İÇERİK FRAME (FORM) =================
        icerik = ttk.Frame(ana)
        icerik.pack(fill="both", expand=True)

        # ================= ALT BUTON FRAME =================
        alt = ttk.Frame(ana)
        alt.pack(fill="x", pady=(10, 0))

        # ================= DEĞİŞKENLER =================
        kategori = tk.StringVar(value="Diğer")
        odeme = tk.StringVar()
        urun = tk.StringVar()
        aciklama = tk.StringVar()
        miktar = tk.StringVar(value="1")
        fiyat = tk.StringVar(value="0")
        toplam = tk.StringVar(value="0.00")

        # ================= FORM =================
        def satir(label, widget, row):
            ttk.Label(icerik, text=label, font=("Segoe UI", 10, "bold"))\
                .grid(row=row, column=0, sticky="w", pady=6)
            widget.grid(row=row, column=1, sticky="ew", pady=6)

        icerik.columnconfigure(1, weight=1)

        satir("Gider Türü",
            ttk.Combobox(icerik, textvariable=kategori,
                        values=["Kira", "Elektrik", "Su", "Diğer"],
                        state="readonly"),
            0)

        satir("Ödeme Türü",
            ttk.Combobox(icerik, textvariable=odeme,
                        values=["Nakit", "Kart", "Havale"],
                        state="readonly"),
            1)

        satir("Açıklama",
            ttk.Entry(icerik, textvariable=aciklama),
            3)

        satir("Miktar",
            ttk.Entry(icerik, textvariable=miktar),
            4)

        satir("Birim Fiyat",
            ttk.Entry(icerik, textvariable=fiyat),
            5)

        satir("Toplam",
            ttk.Entry(icerik, textvariable=toplam, state="readonly"),
            6)

        # ================= TOPLAM HESAP =================
        def toplam_hesapla(*_):
            try:
                t = float(miktar.get()) * float(fiyat.get())
                toplam.set(f"{t:.2f}")
            except:
                toplam.set("0.00")

        miktar.trace_add("write", toplam_hesapla)
        fiyat.trace_add("write", toplam_hesapla)

        # ================= KAYDET =================
        def kaydet_gider():
            try:
                miktar_f = float(miktar.get())
                birim_f = float(fiyat.get())
                tutar = float(toplam.get())
            except:
                messagebox.showerror("Hata", "Miktar veya fiyat hatalı")
                return

            giderler.append({
                "tarih": datetime.now().strftime("%d.%m.%Y %H:%M"),
                "kategori": kategori.get(),
                "urun": urun.get(),
                "aciklama": aciklama.get(),
                "miktar": miktar_f,
                "odeme": odeme.get(),
                "tutar": tutar,
                "kullanici": aktif_kullanici
            })

            # === STOK ARTIR ===
            if kategori.get() == "Malzeme" and urun.get():
                kod = next((k for k, v in urunler.items()
                            if v["ad"] == urun.get()), None)

                if kod:
                    urunler[kod]["stok"] = round(
                        urunler[kod]["stok"] + miktar_f, 2
                    )
                    urunler[kod]["son_alis_fiyat"] = birim_f

                    hareketler.append({
                        "tarih": datetime.now().strftime("%d.%m.%Y %H:%M"),
                        "kullanici": aktif_kullanici,
                        "urun": urun.get(),
                        "miktar": miktar_f,
                        "tur": "GİRİŞ",
                        "aciklama": "Satın alma"
                    })

                    kaydet(DOSYA_URUN, urunler)
                    kaydet(DOSYA_HAREKET, hareketler)

            kaydet(DOSYA_GIDER, giderler)

            messagebox.showinfo("Başarılı", "Gider kaydedildi")
            ana_sayfa_goster()

        # ================= ALT BUTON =================
        ttk.Button(
            alt,
            text="💾 Kaydet",
            command=kaydet_gider
        ).pack(pady=5)



    # ---------- MUHASEBE ----------
    def manuel_gelir_ekle():
        temizle_orta_alan()

        main_frame = ttk.Frame(content_frame, padding=40)
        main_frame.pack(fill="both", expand=True)

        tutar = tk.StringVar()
        aciklama = tk.StringVar()

        form = ttk.Frame(main_frame, padding=30)
        form.place(relx=0.5, rely=0.5, anchor="center")

        label_font = ("Segoe UI", 12, "bold")

        # -------- TUTAR --------
        ttk.Label(
            form,
            text="Gelir Tutarı (₺)",
            font=label_font
        ).grid(row=0, column=0, sticky="w", pady=10)

        ttk.Entry(
            form,
            textvariable=tutar,
            width=30
        ).grid(row=0, column=1, pady=10)

        # -------- AÇIKLAMA --------
        ttk.Label(
            form,
            text="Açıklama",
            font=label_font
        ).grid(row=1, column=0, sticky="w", pady=10)

        ttk.Entry(
            form,
            textvariable=aciklama,
            width=30
        ).grid(row=1, column=1, pady=10)

        # -------- KAYDET --------
        def kaydet_gelir():
            try:
                t = float(tutar.get())
            except:
                messagebox.showerror("Hata", "Geçerli bir tutar girin")
                return

            gelirler.append(
                {
                    "tarih": datetime.now().strftime("%d.%m.%Y %H:%M"),
                    "tutar": t,
                    "aciklama": aciklama.get().strip(),
                    "kullanici": aktif_kullanici,
                }
            )

            kaydet(DOSYA_GELIR, gelirler)
            messagebox.showinfo("Kaydedildi", "Gelir kaydedildi")
            ana_sayfa_goster()

        ttk.Button(
            form,
            text="💾 Kaydet",
            width=22,
            command=kaydet_gelir
        ).grid(row=2, column=0, columnspan=2, pady=25)

    
    def rapor():
        global gelirler, giderler

        gelirler = yukle(DOSYA_GELIR, [])
        giderler = yukle(DOSYA_GIDER, [])

        pdfmetrics.registerFont(UnicodeCIDFont("HeiseiMin-W3"))
        temizle_orta_alan()

        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas

        main = ttk.Frame(content_frame, padding=20)
        main.pack(fill="both", expand=True)

        baslangic = tk.StringVar()
        bitis = tk.StringVar()

    # ---------- SOL TARAF ----------
        sol = ttk.Frame(main)
        sol.pack(side="left", fill="y", padx=20)

        ttk.Label(sol, text="Başlangıç Tarihi").pack(anchor="w")
        DateEntry(sol, textvariable=baslangic, date_pattern="dd.MM.yyyy", width=18)\
            .pack(pady=5)

        ttk.Label(sol, text="Bitiş Tarihi").pack(anchor="w", pady=(10, 0))
        DateEntry(sol, textvariable=bitis, date_pattern="dd.MM.yyyy", width=18)\
            .pack(pady=5)

    # ---------- ÖZET ----------
        sonuc_frame = ttk.Frame(sol)
        sonuc_frame.pack(pady=20, anchor="w")

        font_lbl = ("Segoe UI", 11, "bold")

        ttk.Label(sonuc_frame, text="Toplam Gelir", font=font_lbl).grid(row=0, column=0, sticky="w")
        ttk.Label(sonuc_frame, text=":", font=font_lbl).grid(row=0, column=1, padx=6)
        lbl_gelir = ttk.Label(sonuc_frame, text="0.00 ₺", font=font_lbl)
        lbl_gelir.grid(row=0, column=2, sticky="w")

        ttk.Label(sonuc_frame, text="Toplam Gider", font=font_lbl).grid(row=1, column=0, sticky="w")
        ttk.Label(sonuc_frame, text=":", font=font_lbl).grid(row=1, column=1, padx=6)
        lbl_gider = ttk.Label(sonuc_frame, text="0.00 ₺", font=font_lbl)
        lbl_gider.grid(row=1, column=2, sticky="w")

        ttk.Label(sonuc_frame, text="Net Sonuç", font=font_lbl).grid(row=2, column=0, sticky="w")
        ttk.Label(sonuc_frame, text=":", font=font_lbl).grid(row=2, column=1, padx=6)
        lbl_net = ttk.Label(sonuc_frame, text="0.00 ₺", font=font_lbl)
        lbl_net.grid(row=2, column=2, sticky="w")
        lbl_nakit = ttk.Label(sonuc_frame, text="0.00 ₺", font=font_lbl)
        lbl_kart = ttk.Label(sonuc_frame, text="0.00 ₺", font=font_lbl)

        ttk.Label(sonuc_frame, text="Nakit Toplam", font=font_lbl).grid(row=3, column=0, sticky="w")
        ttk.Label(sonuc_frame, text=":", font=font_lbl).grid(row=3, column=1)
        lbl_nakit.grid(row=3, column=2, sticky="w")

        ttk.Label(sonuc_frame, text="Kart Toplam", font=font_lbl).grid(row=4, column=0, sticky="w")
        ttk.Label(sonuc_frame, text=":", font=font_lbl).grid(row=4, column=1)
        lbl_kart.grid(row=4, column=2, sticky="w")


    # ---------- SAĞ TARAF ----------
        sag = ttk.Frame(main)
        sag.pack(side="left", fill="both", expand=True)

        ttk.Label(sag, text="Gelirler", font=("Segoe UI", 11, "bold")).pack()

        gelir_tree = ttk.Treeview(
            sag,
            columns=("Tarih", "Masa", "Ödeme", "Tutar", "Açıklama", "Kullanıcı"),
            show="headings",
            height=8
        )

        # Başlıklar
        gelir_tree.heading("Tarih", text="Tarih")
        gelir_tree.heading("Masa", text="Masa")
        gelir_tree.heading("Ödeme", text="Ödeme")
        gelir_tree.heading("Tutar", text="Toplam Tutar")
        gelir_tree.heading("Açıklama", text="Açıklama")
        gelir_tree.heading("Kullanıcı", text="Kullanıcı")

        # Genişlik + center
        gelir_tree.column("Tarih", width=140, anchor="center")
        gelir_tree.column("Masa", width=90, anchor="center")
        gelir_tree.column("Ödeme", width=90, anchor="center")
        gelir_tree.column("Tutar", width=120, anchor="center")
        gelir_tree.column("Açıklama", width=180, anchor="center")
        gelir_tree.column("Kullanıcı", width=100, anchor="center")

        gelir_tree.pack(fill="x", pady=5)

        ttk.Label(sag, text="Giderler", font=("Segoe UI", 11, "bold")).pack(pady=(10, 0))

        gider_tree = ttk.Treeview(sag,columns=("Tarih","Kategori","Ürün","Miktar","BirimFiyat","Tutar","Açıklama","Kullanıcı"),
        show="headings",height=8
)
        # ===== GİDER BAŞLIKLARI =====
        gider_tree.heading("Tarih", text="Tarih")
        gider_tree.heading("Kategori", text="Kategori")
        gider_tree.heading("Ürün", text="Ürün")
        gider_tree.heading("Miktar", text="Miktar")
        gider_tree.heading("BirimFiyat", text="Birim Fiyat")
        gider_tree.heading("Tutar", text="Toplam Tutar")
        gider_tree.heading("Açıklama", text="Açıklama")
        gider_tree.heading("Kullanıcı", text="Kullanıcı")
        gider_tree.pack(fill="x", pady=5)
        
        # ===== GİDER KOLON GENİŞLİKLERİ =====
        gider_tree.column("Tarih", width=130, anchor="center")
        gider_tree.column("Kategori", width=110, anchor="center")
        gider_tree.column("Ürün", width=140, anchor="center")
        gider_tree.column("Miktar", width=80, anchor="center")
        gider_tree.column("BirimFiyat", width=110, anchor="center")
        gider_tree.column("Tutar", width=120, anchor="center")
        gider_tree.column("Açıklama", width=160, anchor="center")
        gider_tree.column("Kullanıcı", width=100, anchor="center")

        # ---------- HESAPLA ----------
        def hesapla():
            global gelirler, giderler

            # 🔥 BACKEND’DEN ANLIK GELİRLERİ ÇEK
            gelirler = gelirleri_al()

            # 🔥 GİDERLER HALA DOSYADAN
            giderler = yukle(DOSYA_GIDER, [])

            try:
                b1 = datetime.strptime(baslangic.get(), "%d.%m.%Y")
                b2 = datetime.strptime(bitis.get(), "%d.%m.%Y")
            except:
                messagebox.showerror("Hata", "Tarih formatı GG.AA.YYYY")
                return

            gelir_tree.delete(*gelir_tree.get_children())
            gider_tree.delete(*gider_tree.get_children())

            toplam_gelir = 0
            toplam_gider = 0
            nakit_toplam = 0
            kart_toplam = 0

            # -------- GELİRLER --------
            for g in gelirler:
                try:
                    t = datetime.strptime(g["tarih"][:10], "%d.%m.%Y")
                except:
                    continue

                if b1 <= t <= b2:
                    tutar = float(g.get("tutar", 0))
                    toplam_gelir += tutar

                    if g.get("odeme") == "Nakit":
                        nakit_toplam += tutar
                    elif g.get("odeme") == "Kart":
                        kart_toplam += tutar
                    else:
                        nakit_toplam += tutar  # manuel / bilinmeyen

                    gelir_tree.insert(
                        "",
                        "end",
                        values=(
                            g.get("tarih", ""),
                            g.get("masa", "-"),
                            g.get("odeme", "-"),
                            f"{tutar:.2f}", 
                            g.get("aciklama", ""),
                            g.get("kullanici", "")
                        )
                    )

            # -------- GİDERLER --------
            for g in giderler:
                try:
                    t = datetime.strptime(g["tarih"][:10], "%d.%m.%Y")
                except:
                    continue

                if not (b1 <= t <= b2):
                    continue

                tutar = float(g.get("tutar", 0))
                miktar = float(g.get("miktar", 0))

                birim_fiyat = round(tutar / miktar, 2) if miktar > 0 else 0

                toplam_gider += tutar

                gider_tree.insert(
                    "",
                    "end",
                    values=(
                        g.get("tarih", ""),
                        g.get("kategori", ""),
                        g.get("urun", ""),
                        f"{miktar:.2f}",
                        f"{birim_fiyat:.2f}",
                        f"{tutar:.2f}",
                        g.get("aciklama", ""),
                        g.get("kullanici", "")
                    )
                )

            # -------- LABEL GÜNCELLE --------
            lbl_gelir.config(text=f"{toplam_gelir:.2f} ₺")
            lbl_gider.config(text=f"{toplam_gider:.2f} ₺")
            lbl_net.config(text=f"{(toplam_gelir - toplam_gider):.2f} ₺")
            lbl_nakit.config(text=f"{nakit_toplam:.2f} ₺")
            lbl_kart.config(text=f"{kart_toplam:.2f} ₺")

    # ---------- PDF ----------
        def pdf_aktar():
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.cidfonts import UnicodeCIDFont
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas

            pdfmetrics.registerFont(UnicodeCIDFont("HeiseiMin-W3"))

            dosya = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF Dosyası", "*.pdf")],
                title="Raporu PDF olarak kaydet"
            )
            if not dosya:
                return

            c = canvas.Canvas(dosya, pagesize=A4)
            w, h = A4
            y = h - 40

    # ---- BAŞLIK ----
            c.setFont("HeiseiMin-W3", 14)
            c.drawString(40, y, "Kâr / Zarar Raporu")
            y -= 30

    # ---- GELİRLER ----
            c.setFont("HeiseiMin-W3", 12)
            c.drawString(40, y, "Gelirler")
            y -= 20

            c.setFont("HeiseiMin-W3", 9)
            for item in gelir_tree.get_children():
                t = gelir_tree.item(item)["values"]
                c.drawString(40, y, f"{t[0]} | {t[1]} | {t[2]} ₺ | {t[3]}")
                y -= 14
                if y < 40:
                    c.showPage()
                    y = h - 40
                    c.setFont("HeiseiMin-W3", 9)

    # ---- GİDERLER ----
            y -= 20
            c.setFont("HeiseiMin-W3", 12)
            c.drawString(40, y, "Giderler")
            y -= 20

            c.setFont("HeiseiMin-W3", 9)
            for item in gider_tree.get_children():
                t = gider_tree.item(item)["values"]
                c.drawString(40, y, f"{t[0]} | {t[1]} | {t[2]} | {t[3]} ₺")
                y -= 14
                if y < 40:
                    c.showPage()
                    y = h - 40
                    c.setFont("HeiseiMin-W3", 9)

            c.save()
            messagebox.showinfo("PDF Oluşturuldu", "PDF Türkçe karakterlerle oluşturuldu.")

        def excel_aktar():
            dosya = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Dosyası", "*.xlsx")],
                title="Raporu Excel olarak kaydet"
            )
            if not dosya:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Kar-Zarar"

            bold = Font(bold=True)

            # ===== ÖZET =====
            ws["A1"] = "KÂR / ZARAR RAPORU"
            ws["A1"].font = bold

            ws["A3"] = "Toplam Gelir"
            ws["B3"] = lbl_gelir.cget("text")

            ws["A4"] = "Toplam Gider"
            ws["B4"] = lbl_gider.cget("text")

            ws["A5"] = "Net Sonuç"
            ws["B5"] = lbl_net.cget("text")

            ws["A6"] = "Nakit Toplam"
            ws["B6"] = lbl_nakit.cget("text")

            ws["A7"] = "Kart Toplam"
            ws["B7"] = lbl_kart.cget("text")

            for i in range(3, 8):
                ws[f"A{i}"].font = bold

            # ===== GELİRLER =====
            row = 9
            ws[f"A{row}"] = "GELİRLER"
            ws[f"A{row}"].font = bold
            row += 1

            headers = ["Tarih", "Masa", "Ödeme", "Tutar", "Kullanıcı"]
            for col, h in enumerate(headers, start=1):
                ws.cell(row=row, column=col, value=h).font = bold
            row += 1

            for item in gelir_tree.get_children():
                for col, val in enumerate(gelir_tree.item(item)["values"], start=1):
                    ws.cell(row=row, column=col, value=val)
                row += 1

            # ===== GİDERLER =====
            row += 2
            ws[f"A{row}"] = "GİDERLER"
            ws[f"A{row}"].font = bold
            row += 1

            headers = ["Tarih", "Kategori", "Ürün", "Açıklama", "Tutar", "Kullanıcı"]
            for col, h in enumerate(headers, start=1):
                ws.cell(row=row, column=col, value=h).font = bold
            row += 1

            for item in gider_tree.get_children():
                for col, val in enumerate(gider_tree.item(item)["values"], start=1):
                    ws.cell(row=row, column=col, value=val)
                row += 1

            # Otomatik kolon genişliği
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 18

            wb.save(dosya)
            messagebox.showinfo("Excel Oluşturuldu", "Excel raporu başarıyla oluşturuldu.")


    # ---------- BUTONLAR ----------
        ttk.Button(sol, text="📊 Hesapla", command=hesapla).pack(pady=(10, 5))
        ttk.Button(sol, text="📄 PDF'e Aktar", command=pdf_aktar).pack(pady=5)
        ttk.Button(sol, text="📊 Excel'e Aktar", command=excel_aktar).pack()

    menubar = tk.Menu(root)
    root.config(menu=menubar)

    # ===== RESTORAN =====
    if "Restoran" in aktif_yetkiler and any(aktif_yetkiler["Restoran"].values()):
        restoran = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="🍽 Restoran", menu=restoran)

        if aktif_yetkiler["Restoran"].get("Dashboard"):
            restoran.add_command(label="📊 Dashboard", command=dashboard)

        if aktif_yetkiler["Restoran"].get("Menü Yönetimi"):
            restoran.add_command(label="🍽 Menü Yönetimi", command=menu_yonetimi)

        if aktif_yetkiler["Restoran"].get("Adisyonlar"):
            restoran.add_command(label="🧾 Adisyonlar", command=adisyon_ekrani)

    # ===== DEPO =====
    if yetkisi_var("Depo"):
        depo = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="📦 Depo", menu=depo)

        if yetkisi_var("Depo", "Stok Görüntüle"):
            depo.add_command(label="📋 Stok Görüntüle", command=stok_goster)

        if yetkisi_var("Depo", "Stok Çıkış"):
            depo.add_command(label="📤 Stok Çıkış", command=lambda: stok_hareket("ÇIKIŞ"))

        if yetkisi_var("Depo", "Stok Hareketleri"):
            depo.add_command(label="🔄 Stok Hareketleri", command=stok_hareketleri)

        if yetkisi_var("Depo", "Ürün Tanımlama"):
            depo.add_command(label="🏷 Ürün Tanımlama", command=urun_yonetimi)


    # ===== SATIN ALMA =====
        if yetkisi_var("Satın Alma"):
            satin = tk.Menu(menubar, tearoff=0)
            menubar.add_cascade(label="🛒 Satın Alma", menu=satin)

            if yetkisi_var("Satın Alma", "Gider Girişi"):
                satin.add_command(label="💸 Gider Girişi", command=gider_giris_ekrani)

    # ===== MUHASEBE =====
        if yetkisi_var("Muhasebe"):
            muh = tk.Menu(menubar, tearoff=0)
            menubar.add_cascade(label="💰 Muhasebe", menu=muh)

        if yetkisi_var("Muhasebe", "Gelir Girişi"):
            muh.add_command(label="💵 Gelir Girişi", command=manuel_gelir_ekle)

        if yetkisi_var("Muhasebe", "Kâr / Zarar"):
            muh.add_command(label="📊 Kâr / Zarar", command=rapor)
            

    
    # ===== RAPOR =====
        if yetkisi_var("Rapor"):
            rapor_menu = tk.Menu(menubar, tearoff=0)
            menubar.add_cascade(label="📈 Rapor", menu=rapor_menu)

        if yetkisi_var("Rapor", "Satış Raporu"):
            rapor_menu.add_command(
                label="📈 Satış Raporu",
                command=lambda: satis_raporu(content_frame, DOSYA_GELIR)
            )

        if yetkisi_var("Rapor", "Ödeme Raporu"):
            rapor_menu.add_command(
                label="💳 Ödeme Raporu",
                command=lambda: odeme_raporu(content_frame)
            )
        if yetkisi_var("Rapor", "Ürün Satış Raporu"):
            rapor_menu.add_command(
                label="📦 Ürün Satış Raporu",
                command=lambda: urun_satis_adet_raporu(content_frame, DOSYA_SATIS_DETAY)
            )

        if yetkisi_var("Rapor", "Satın Alma Fiyat Raporu"):
            rapor_menu.add_command(
                label="🛒 Satın Alma Birim Fiyat Raporu",
                command=lambda: satin_alma_fiyat_raporu(content_frame, DOSYA_GIDER)
            )
        if yetkisi_var("Rapor", "Kâr / Zarar (Yeni)"):
            rapor_menu.add_command(
                label="📊 Kâr / Zarar (Yeni)",
                command=lambda: kar_zarar_yeni(content_frame, temizle_orta_alan)
            )


    # ===== ADMIN (sadece admin kullanıcı görür) =====
    if aktif_rol == "admin":
        admin = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="👤 Admin", menu=admin)

        admin.add_command(
            label="👥 Kullanıcı Yönetimi",
            command=kullanici_yonetimi
        )

        admin.add_separator()

        admin.add_command(
            label="🔄 Sistem Güncelleme",
            command=guncelleme_baslat
        )



    # ===== YEDEKLEME =====
        yedek_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="💾 Yedekleme", menu=yedek_menu)

        yedek_menu.add_command(label="📥 Yedek Al", command=yedek_al)

        if aktif_rol == "admin":
            yedek_menu.add_command(
                label="📤 Yedekten Geri Yükle",
                command=yedekten_yukle
            )

def yedek_al():
    try:
        yedek_klasor = os.path.join(
            BASE_DIR, "yedekler", datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        )

        os.makedirs(yedek_klasor, exist_ok=True)

        for dosya in [DOSYA_URUN, DOSYA_HAREKET, DOSYA_GIDER, DOSYA_GELIR, DOSYA_USER]:
            if os.path.exists(dosya):
                shutil.copy(dosya, yedek_klasor)

        print("✔ Otomatik yedek alındı:", yedek_klasor)

    except Exception as e:
        print("❌ Yedekleme hatası:", e)


def yedekten_yukle():
    if aktif_rol != "admin":
        messagebox.showerror("Yetki", "Sadece admin yedekten geri yükleyebilir.")
        return

    secilen_klasor = filedialog.askdirectory(
        title="Yedek Klasörü Seç", initialdir=os.path.join(BASE_DIR, "yedekler")
    )
    if not secilen_klasor:
        return

    if not messagebox.askyesno(
        "Yedekten Geri Yükle",
        "Bu işlem tüm mevcut verilerin ÜZERİNE yazacak!\nDevam edilsin mi?",
    ):
        return

    try:
        dosyalar = {
            "products.json": DOSYA_URUN,
            "hareketler.json": DOSYA_HAREKET,
            "expenses.json": DOSYA_GIDER,
            "revenue.json": DOSYA_GELIR,
            "users.json": DOSYA_USER,
        }

        for f, hedef in dosyalar.items():
            yol = os.path.join(secilen_klasor, f)
            if os.path.exists(yol):
                shutil.copy(yol, hedef)

        messagebox.showinfo("Geri Yükleme", "Yedek yüklendi.\nProgram kapanacak.")
        root.destroy()

    except Exception as e:
        messagebox.showerror("Hata", str(e))


def program_kapanirken():
    yedek_al()
    root.destroy()

    # 🔔 Kritik stok uyarısı

def kritik_kontrol():
    kritikler = [
        f'{u["ad"]} → {u["stok"]} {u["birim"]} (kritik: {u.get("kritik", 0)})'
        for u in urunler.values()
        if u["stok"] <= u.get("kritik", 0)
    ]

    if kritikler:
        messagebox.showwarning(
            "⚠️ Kritik Stok Uyarısı",
            "Aşağıdaki ürünler kritik seviyede:\n\n" + "\n".join(kritikler),
        )

# ================= PROGRAM =================
login_ekrani()
root.protocol("WM_DELETE_WINDOW", program_kapanirken)
root.mainloop()

