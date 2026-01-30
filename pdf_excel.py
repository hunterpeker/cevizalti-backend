from tkinter import filedialog, messagebox
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Font


# ==================================================
# TREEVIEW -> PDF
# ==================================================
def treeview_pdf_aktar(baslik, treeviews):
    # Türkçe karakter destekli font
    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiMin-W3"))

    dosya = filedialog.asksaveasfilename(
        title="PDF Kaydet",
        defaultextension=".pdf",
        filetypes=[("PDF Dosyası", "*.pdf")]
    )
    if not dosya:
        return

    c = canvas.Canvas(dosya, pagesize=A4)
    sayfa_genislik, sayfa_yukseklik = A4

    # ---------- AYARLAR ----------
    baslangic_x = 40
    baslangic_y = sayfa_yukseklik - 60
    satir_yukseklik = 18
    font_boyut = 10

    # ---------- BAŞLIK ----------
    c.setFont("HeiseiMin-W3", 16)
    c.drawString(baslangic_x, baslangic_y, baslik)
    y = baslangic_y - 30

    c.setFont("HeiseiMin-W3", font_boyut)

    for tree in treeviews:
        kolonlar = tree["columns"]

        # Kolon genişlikleri (ihtiyaca göre artır/azalt)
        kolon_genislik = {
            "Tarih": 95,
            "Masa": 60,
            "Ödeme": 70,
            "Tutar": 70,
            "Kullanıcı": 80
        }

        # ---------- BAŞLIKLAR ----------
        x = baslangic_x
        for kolon in kolonlar:
            c.drawString(x, y, kolon)
            x += kolon_genislik.get(kolon, 70)

        y -= satir_yukseklik

        # ---------- SATIRLAR ----------
        for item in tree.get_children():
            x = baslangic_x
            degerler = tree.item(item)["values"]

            for kolon, deger in zip(kolonlar, degerler):
                c.drawString(x, y, str(deger))
                x += kolon_genislik.get(kolon, 70)

            y -= satir_yukseklik

            # Sayfa sonu kontrolü
            if y < 50:
                c.showPage()
                c.setFont("HeiseiMin-W3", font_boyut)
                y = sayfa_yukseklik - 50

        y -= 25  # tablolar arası boşluk

    c.save()
    messagebox.showinfo("PDF", "PDF başarıyla oluşturuldu.")


# ==================================================
# TREEVIEW -> EXCEL
# ==================================================
def treeview_excel_aktar(baslik, treeviews):
    dosya = filedialog.asksaveasfilename(
        title="Excel Kaydet",
        defaultextension=".xlsx",
        filetypes=[("Excel Dosyası", "*.xlsx")]
    )
    if not dosya:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Rapor"

    kalin = Font(bold=True)

    # ---------- BAŞLIK ----------
    ws["A1"] = baslik
    ws["A1"].font = kalin

    satir = 3

    for tree in treeviews:
        kolonlar = tree["columns"]

        # Kolon başlıkları
        for i, kolon in enumerate(kolonlar, start=1):
            hucre = ws.cell(row=satir, column=i, value=kolon)
            hucre.font = kalin

        satir += 1

        # Satırlar
        for item in tree.get_children():
            degerler = tree.item(item)["values"]
            for i, deger in enumerate(degerler, start=1):
                ws.cell(row=satir, column=i, value=deger)
            satir += 1

        satir += 2  # tablolar arası boşluk

    # Otomatik kolon genişliği
    for kolon in ws.columns:
        ws.column_dimensions[kolon[0].column_letter].width = 18

    wb.save(dosya)
    messagebox.showinfo("Excel", "Excel başarıyla oluşturuldu.")
